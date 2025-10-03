import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from moviepy.editor import VideoFileClip
from scipy.io import wavfile
from scipy.fft import fft, fftfreq
from scipy import signal
import os
import threading
from pathlib import Path
import platform
from PIL import Image
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart

# 優化的色彩配置 - 增強邊框和分隔效果
JAPANESE_COLORS = {
    # 主要背景色 - 溫暖的淺米色
    'bg_primary': '#F7F3E9',
    # 次要背景色 - 更淺的米白色
    'bg_secondary': '#FEFCF7',
    # 卡片背景色 - 純白色
    'bg_card': '#FFFFFF',
    # 主要文字色 - 黑色（易讀性）
    'text_primary': '#000000',
    # 次要文字色 - 深灰色
    'text_secondary': '#333333',
    # 提示文字色 - 中灰色
    'text_hint': '#666666',
    # 版本資訊文字色 - 淺灰色
    'text_version': '#999999',
    # 主要強調色 - 柔和的藍色
    'accent_primary': '#6FA8DC',
    # 次要強調色 - 溫暖的綠色
    'accent_secondary': '#81C784',
    # 警告色 - 柔和的橙色
    'warning': '#FFB74D',
    # 錯誤色 - 柔和的紅色
    'error': '#E57373',
    # 成功色 - 清新的綠色
    'success': '#A5D6A7',
    # 邊框色 - 較深的灰色（增強視覺分隔）
    'border': '#D0D0D0',
    # 分隔線色 - 中等灰色
    'divider': '#E8E8E8',
    # 陰影色 - 淡灰色
    'shadow': '#C0C0C0',
    # 強邊框色 - 用於重要區塊
    'border_strong': '#B0B0B0',
    # 輕邊框色 - 用於次要區塊
    'border_light': '#E0E0E0'
}

# 設置matplotlib日系風格
def setup_matplotlib_japanese_style():
    """設置matplotlib的日系風格"""
    plt.rcParams.update({
        'figure.facecolor': JAPANESE_COLORS['bg_card'],
        'axes.facecolor': JAPANESE_COLORS['bg_secondary'],
        'axes.edgecolor': JAPANESE_COLORS['border'],
        'axes.labelcolor': JAPANESE_COLORS['text_primary'],
        'axes.axisbelow': True,
        'axes.grid': True,
        'grid.color': JAPANESE_COLORS['divider'],
        'grid.linewidth': 0.5,
        'text.color': JAPANESE_COLORS['text_primary'],
        'xtick.color': JAPANESE_COLORS['text_secondary'],
        'ytick.color': JAPANESE_COLORS['text_secondary'],
        'xtick.labelsize': 9,
        'ytick.labelsize': 9,
        'axes.titlesize': 12,
        'axes.labelsize': 10,
        'lines.linewidth': 1.5,
        'patch.linewidth': 0.5,
        'patch.facecolor': JAPANESE_COLORS['accent_primary'],
        'patch.edgecolor': JAPANESE_COLORS['border']
    })

# 設置中文字體
def setup_chinese_font():
    """設置中文字體支援"""
    import matplotlib.font_manager as fm

    # 根據作業系統選擇合適的中文字體
    system = platform.system()
    if system == "Windows":
        chinese_fonts = ['Microsoft YaHei', 'SimHei', 'SimSun', 'KaiTi']
    elif system == "Darwin":  # macOS
        chinese_fonts = ['Hei', 'STHeiti', 'Arial Unicode MS']
    else:  # Linux
        chinese_fonts = ['WenQuanYi Micro Hei', 'Droid Sans Fallback', 'DejaVu Sans']

    # 嘗試設置中文字體
    for font_name in chinese_fonts:
        try:
            plt.rcParams['font.sans-serif'] = [font_name]
            plt.rcParams['axes.unicode_minus'] = False
            # 測試字體是否可用
            fig, ax = plt.subplots(figsize=(1, 1))
            ax.text(0.5, 0.5, '測試', fontsize=12)
            plt.close(fig)
            print(f"成功設置字體: {font_name}")
            return font_name
        except:
            continue

    # 如果都失敗，使用默認字體並警告用戶
    print("警告: 無法找到合適的中文字體，可能會出現顯示問題")
    return None

class VideoAudioAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("Audio Spectrum Analyzer - 影片音訊頻譜分析器")
        self.root.geometry("1440x980")

        # 設置日系風格
        self.setup_japanese_ui_style()

        # 設置matplotlib樣式
        setup_matplotlib_japanese_style()

        # 設置中文字體
        self.chinese_font = setup_chinese_font()

        # 設置tkinter字體
        self.setup_fonts()

        # 變數
        self.video_file = None
        self.temp_audio_file = "temp_audio.wav"
        self.is_processing = False
        self.audio_duration = 0
        self.sample_rate = 0
        self.save_directory = None
        self.captured_screenshots = []

        # 分析類型設定
        self.analysis_type = tk.StringVar(value="high_freq")  # 預設為高頻分析

        # 自訂頻率範圍
        self.custom_freq_start = tk.IntVar(value=1000)
        self.custom_freq_end = tk.IntVar(value=5000)

        # 時間範圍設定
        self.use_time_range = tk.BooleanVar(value=False)
        self.time_start = tk.DoubleVar(value=0.0)
        self.time_end = tk.DoubleVar(value=0.0)

        # 振幅閾值設定
        self.use_amplitude_threshold = tk.BooleanVar(value=False)
        self.amplitude_threshold = tk.IntVar(value=-40)

        # 🆕 多頻帶分析設定
        self.use_multi_band = tk.BooleanVar(value=False)
        self.frequency_bands = []  # 儲存多個頻率範圍

        # 🆕 音訊節錄設定
        self.enable_audio_extract = tk.BooleanVar(value=True)
        self.audio_duration_before = tk.DoubleVar(value=1.0)  # 峰值前幾秒
        self.audio_duration_after = tk.DoubleVar(value=1.0)   # 峰值後幾秒
        self.extracted_audio_clips = []  # 儲存節錄的音訊片段

        # 時頻分析結果
        self.time_freq_data = None
        self.time_peaks = []
        self.multi_band_peaks = {}  # 儲存各頻帶的峰值資料

        # 創建GUI元素
        self.create_widgets()

    def setup_japanese_ui_style(self):
        """設置日系UI風格 - 增強邊框效果"""
        # 設置主窗口背景
        self.root.configure(bg=JAPANESE_COLORS['bg_primary'])

        # 創建自定義樣式
        self.style = ttk.Style()

        # 設置主題
        self.style.theme_use('clam')

        # 配置各種元件的樣式
        self.style.configure('Title.TLabel',
                           font=('Yu Gothic UI', 16, 'bold'),
                           foreground=JAPANESE_COLORS['text_primary'],
                           background=JAPANESE_COLORS['bg_card'])

        self.style.configure('Version.TLabel',
                           font=('Yu Gothic UI', 8),
                           foreground=JAPANESE_COLORS['text_version'],
                           background=JAPANESE_COLORS['bg_card'])

        self.style.configure('Heading.TLabel',
                           font=('Yu Gothic UI', 12, 'bold'),
                           foreground=JAPANESE_COLORS['text_primary'],
                           background=JAPANESE_COLORS['bg_card'])

        self.style.configure('Body.TLabel',
                           font=('Yu Gothic UI', 9),
                           foreground=JAPANESE_COLORS['text_primary'],
                           background=JAPANESE_COLORS['bg_card'])

        self.style.configure('Secondary.TLabel',
                           font=('Yu Gothic UI', 9),
                           foreground=JAPANESE_COLORS['text_secondary'],
                           background=JAPANESE_COLORS['bg_card'])

        self.style.configure('Hint.TLabel',
                           font=('Yu Gothic UI', 8),
                           foreground=JAPANESE_COLORS['text_hint'],
                           background=JAPANESE_COLORS['bg_card'])

        # 按鈕樣式 - 增強邊框
        self.style.configure('Primary.TButton',
                           font=('Yu Gothic UI', 9, 'bold'),
                           foreground='white',
                           background=JAPANESE_COLORS['accent_primary'],
                           borderwidth=2,
                           focuscolor='none',
                           relief='solid')

        self.style.map('Primary.TButton',
                      background=[('active', '#5A9BD4'),
                                ('pressed', '#4A8BC2')],
                      bordercolor=[('active', JAPANESE_COLORS['border_strong'])])

        self.style.configure('Secondary.TButton',
                           font=('Yu Gothic UI', 9),
                           foreground=JAPANESE_COLORS['text_primary'],
                           background=JAPANESE_COLORS['bg_secondary'],
                           borderwidth=2,
                           focuscolor='none',
                           relief='solid')

        self.style.map('Secondary.TButton',
                      background=[('active', JAPANESE_COLORS['divider']),
                                ('pressed', JAPANESE_COLORS['border'])],
                      bordercolor=[('active', JAPANESE_COLORS['border_strong'])])

        # 框架樣式 - 增強邊框
        self.style.configure('Card.TLabelframe',
                           background=JAPANESE_COLORS['bg_card'],
                           borderwidth=2,
                           relief='solid',
                           bordercolor=JAPANESE_COLORS['border'])

        self.style.configure('Card.TLabelframe.Label',
                           font=('Yu Gothic UI', 10, 'bold'),
                           foreground=JAPANESE_COLORS['text_primary'],
                           background=JAPANESE_COLORS['bg_card'])

        # 強調框架樣式 - 更明顯的邊框
        self.style.configure('StrongCard.TLabelframe',
                           background=JAPANESE_COLORS['bg_card'],
                           borderwidth=3,
                           relief='solid',
                           bordercolor=JAPANESE_COLORS['border_strong'])

        self.style.configure('StrongCard.TLabelframe.Label',
                           font=('Yu Gothic UI', 10, 'bold'),
                           foreground=JAPANESE_COLORS['text_primary'],
                           background=JAPANESE_COLORS['bg_card'])

        # 進度條樣式
        self.style.configure('Japanese.Horizontal.TProgressbar',
                           background=JAPANESE_COLORS['accent_primary'],
                           troughcolor=JAPANESE_COLORS['divider'],
                           borderwidth=1,
                           lightcolor=JAPANESE_COLORS['accent_primary'],
                           darkcolor=JAPANESE_COLORS['accent_primary'])

        # Notebook樣式 - 增強邊框
        self.style.configure('Japanese.TNotebook',
                           background=JAPANESE_COLORS['bg_card'],
                           borderwidth=2,
                           relief='solid')

        self.style.configure('Japanese.TNotebook.Tab',
                           background=JAPANESE_COLORS['bg_secondary'],
                           foreground=JAPANESE_COLORS['text_primary'],
                           padding=[20, 8],
                           font=('Yu Gothic UI', 9),
                           borderwidth=1)

        self.style.map('Japanese.TNotebook.Tab',
                      background=[('selected', JAPANESE_COLORS['bg_card']),
                                ('active', JAPANESE_COLORS['divider'])],
                      foreground=[('selected', JAPANESE_COLORS['text_primary'])])

    def setup_fonts(self):
        """設置字體"""
        system = platform.system()
        if system == "Windows":
            self.font_family = 'Yu Gothic UI'
        elif system == "Darwin":  # macOS
            self.font_family = 'Hiragino Sans'
        else:  # Linux
            self.font_family = 'Noto Sans CJK JP'

        # 嘗試使用日系字體，失敗則使用系統默認
        try:
            test_font = (self.font_family, 9)
            test_label = tk.Label(self.root, text="テスト", font=test_font)
            test_label.destroy()
        except:
            self.font_family = 'TkDefaultFont'

        self.fonts = {
            'title': (self.font_family, 16, 'bold'),
            'heading': (self.font_family, 12, 'bold'),
            'body': (self.font_family, 9),
            'small': (self.font_family, 8),
            'code': ('Consolas', 9)
        }

    def create_widgets(self):
        # 創建主容器框架 - 添加明顯邊框
        main_container = tk.Frame(self.root, bg=JAPANESE_COLORS['bg_primary'],
                                 relief='solid', borderwidth=2,
                                 highlightbackground=JAPANESE_COLORS['border_strong'])
        main_container.pack(fill="both", expand=True, padx=10, pady=10)

        # 主框架 - 使用Canvas來實現更好的滾動效果
        main_canvas = tk.Canvas(main_container, bg=JAPANESE_COLORS['bg_primary'],
                               highlightthickness=0,
                               relief='solid', borderwidth=1,
                               highlightbackground=JAPANESE_COLORS['border'])
        main_scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=main_canvas.yview)

        # 創建可滾動的內容框架 - 添加邊框
        self.main_frame = tk.Frame(main_canvas, bg=JAPANESE_COLORS['bg_primary'],
                                  relief='flat', borderwidth=0)

        self.main_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        main_canvas.create_window((0, 0), window=self.main_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=main_scrollbar.set)

        main_canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        main_scrollbar.pack(side="right", fill="y", padx=(5, 5), pady=5)

        # 配置網格權重
        self.main_frame.columnconfigure(0, weight=1)

        # 標題區域
        self.create_title_section()

        # 檔案選擇區域
        self.create_file_selection_section()

        # 分析設定區域
        self.create_analysis_settings_section()

        # 處理控制區域
        self.create_control_section()

        # 截圖預覽區域
        self.create_preview_section()

        # 結果顯示區域
        self.create_results_section()

        # 綁定滾輪事件
        self.bind_mousewheel(main_canvas)

    def bind_mousewheel(self, canvas):
        """綁定滾輪事件"""
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def create_title_section(self):
        """創建標題區域 - 增強邊框"""
        title_frame = ttk.Frame(self.main_frame, style='StrongCard.TLabelframe', padding="25")
        title_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(10, 20), padx=10)
        title_frame.columnconfigure(0, weight=1)

        # 添加陰影效果的內部框架
        shadow_frame = tk.Frame(title_frame, bg=JAPANESE_COLORS['shadow'], height=2)
        shadow_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(15, 0))

        # 標題容器 - 包含主標題和版本資訊
        title_container = ttk.Frame(title_frame)
        title_container.grid(row=0, column=0, sticky=(tk.W, tk.E))
        title_container.columnconfigure(0, weight=1)

        # 主標題
        title_label = ttk.Label(title_container,
                              text="🎵 Audio Spectrum Analyzer",
                              style='Title.TLabel')
        title_label.grid(row=0, column=0)

        # 版本資訊 - 右上角，增加邊框
        version_frame = tk.Frame(title_container, bg=JAPANESE_COLORS['bg_secondary'],
                               relief='solid', borderwidth=1,
                               highlightbackground=JAPANESE_COLORS['border'])
        version_frame.grid(row=0, column=1, sticky=tk.E, padx=(20, 0))

        version_label = tk.Label(version_frame,
                               text="作成：設計分室　Ver.1.0",
                               font=self.fonts['small'],
                               fg=JAPANESE_COLORS['text_version'],
                               bg=JAPANESE_COLORS['bg_secondary'])
        version_label.pack(padx=8, pady=4)

        # 副標題
        subtitle_label = ttk.Label(title_frame,
                                 text="影片音訊頻譜分析器 - 含時間分析及智能截圖功能",
                                 style='Body.TLabel')
        subtitle_label.grid(row=1, column=0, pady=(10, 0))

    def create_file_selection_section(self):
        """創建檔案選擇區域 - 增強視覺分隔"""
        file_frame = ttk.LabelFrame(self.main_frame, text="📁 步驟 1: 選擇檔案與設定",
                                   style='StrongCard.TLabelframe', padding="25")
        file_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 20), padx=10)
        file_frame.columnconfigure(1, weight=1)

        # 影片檔案選擇區域 - 添加內部邊框
        video_section = tk.Frame(file_frame, bg=JAPANESE_COLORS['bg_card'],
                               relief='solid', borderwidth=1,
                               highlightbackground=JAPANESE_COLORS['border_light'])
        video_section.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15), padx=5)
        video_section.columnconfigure(1, weight=1)

        video_label = tk.Label(video_section, text="影片檔案:",
                             font=self.fonts['heading'],
                             fg=JAPANESE_COLORS['text_primary'],
                             bg=JAPANESE_COLORS['bg_card'])
        video_label.grid(row=0, column=0, sticky=tk.W, pady=15, padx=15)

        file_select_frame = tk.Frame(video_section, bg=JAPANESE_COLORS['bg_card'])
        file_select_frame.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=15, padx=15)
        file_select_frame.columnconfigure(1, weight=1)

        self.select_button = ttk.Button(file_select_frame, text="選擇影片檔案",
                                       command=self.select_file, style='Primary.TButton')
        self.select_button.grid(row=0, column=0, padx=(0, 10))

        self.file_label = tk.Label(file_select_frame, text="尚未選擇檔案",
                                 font=self.fonts['body'],
                                 fg=JAPANESE_COLORS['text_hint'],
                                 bg=JAPANESE_COLORS['bg_card'])
        self.file_label.grid(row=0, column=1, sticky=(tk.W, tk.E))

        # 支援格式說明區域 - 添加邊框
        format_section = tk.Frame(file_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                relief='solid', borderwidth=1,
                                highlightbackground=JAPANESE_COLORS['border_light'])
        format_section.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15), padx=5)

        format_title = tk.Label(format_section, text="📋 支援的影片格式:",
                              font=self.fonts['body'],
                              fg=JAPANESE_COLORS['text_secondary'],
                              bg=JAPANESE_COLORS['bg_secondary'])
        format_title.grid(row=0, column=0, sticky=tk.W, padx=15, pady=(10, 5))

        format_list = tk.Label(format_section,
                             text="MP4, AVI, MOV, MKV, WMV, FLV, WebM, M4V, 3GP, MPG/MPEG",
                             font=self.fonts['body'],
                             fg=JAPANESE_COLORS['text_primary'],
                             bg=JAPANESE_COLORS['bg_secondary'])
        format_list.grid(row=1, column=0, sticky=tk.W, padx=30, pady=(0, 10))

        # 檔案限制說明區域 - 添加邊框
        limitation_section = tk.Frame(file_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                    relief='solid', borderwidth=1,
                                    highlightbackground=JAPANESE_COLORS['border_light'])
        limitation_section.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15), padx=5)

        limitation_title = tk.Label(limitation_section, text="⚠️ 檔案來源限制:",
                                   font=self.fonts['body'],
                                   fg=JAPANESE_COLORS['text_secondary'],
                                   bg=JAPANESE_COLORS['bg_secondary'])
        limitation_title.grid(row=0, column=0, sticky=tk.W, padx=15, pady=(10, 5))

        limitations = [
            "• 建議檔案大小小於 2GB，以獲得最佳處理效能",
            "• 不支援加密或受保護的影片檔案",
            "• 音訊取樣率建議 44.1kHz 或更高，以確保分析精度",
            "• 影片長度超過 30 分鐘時，處理時間會較長"
        ]

        for i, limitation in enumerate(limitations):
            limit_label = tk.Label(limitation_section, text=limitation,
                                 font=self.fonts['body'],
                                 fg=JAPANESE_COLORS['text_primary'],
                                 bg=JAPANESE_COLORS['bg_secondary'])
            limit_label.grid(row=i+1, column=0, sticky=tk.W, padx=30, pady=2)

        # 在最後添加底部間距
        bottom_spacer = tk.Frame(limitation_section, bg=JAPANESE_COLORS['bg_secondary'], height=10)
        bottom_spacer.grid(row=len(limitations)+1, column=0, sticky=(tk.W, tk.E))

        # 儲存位置和截圖設定區域 - 並排顯示並添加邊框
        settings_container = tk.Frame(file_frame, bg=JAPANESE_COLORS['bg_card'])
        settings_container.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10), padx=5)
        settings_container.columnconfigure(0, weight=1)
        settings_container.columnconfigure(1, weight=1)

        # 儲存位置區域
        save_section = tk.Frame(settings_container, bg=JAPANESE_COLORS['bg_secondary'],
                              relief='solid', borderwidth=1,
                              highlightbackground=JAPANESE_COLORS['border_light'])
        save_section.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5), pady=5)
        save_section.columnconfigure(0, weight=1)

        save_title = tk.Label(save_section, text="💾 儲存位置",
                            font=self.fonts['heading'],
                            fg=JAPANESE_COLORS['text_primary'],
                            bg=JAPANESE_COLORS['bg_secondary'])
        save_title.grid(row=0, column=0, pady=(10, 5), padx=15)

        save_button_frame = tk.Frame(save_section, bg=JAPANESE_COLORS['bg_secondary'])
        save_button_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=15, pady=(0, 5))
        save_button_frame.columnconfigure(0, weight=1)

        self.save_dir_button = ttk.Button(save_button_frame, text="選擇儲存位置",
                                         command=self.select_save_directory, style='Secondary.TButton')
        self.save_dir_button.grid(row=0, column=0, pady=5)

        self.save_dir_label = tk.Label(save_section, text="將儲存到程式所在目錄",
                                     font=self.fonts['small'],
                                     fg=JAPANESE_COLORS['text_hint'],
                                     bg=JAPANESE_COLORS['bg_secondary'],
                                     wraplength=200)
        self.save_dir_label.grid(row=2, column=0, padx=15, pady=(0, 10))

        # 截圖設定區域
        screenshot_section = tk.Frame(settings_container, bg=JAPANESE_COLORS['bg_secondary'],
                                    relief='solid', borderwidth=1,
                                    highlightbackground=JAPANESE_COLORS['border_light'])
        screenshot_section.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0), pady=5)

        screenshot_title = tk.Label(screenshot_section, text="📸 截圖設定",
                                   font=self.fonts['heading'],
                                   fg=JAPANESE_COLORS['text_primary'],
                                   bg=JAPANESE_COLORS['bg_secondary'])
        screenshot_title.grid(row=0, column=0, columnspan=3, pady=(10, 5), padx=15)

        screenshot_config_frame = tk.Frame(screenshot_section, bg=JAPANESE_COLORS['bg_secondary'])
        screenshot_config_frame.grid(row=1, column=0, columnspan=3, padx=15, pady=(0, 5))

        tk.Label(screenshot_config_frame, text="截圖數量:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=0, sticky=tk.W)

        self.screenshot_count = tk.IntVar(value=5)
        screenshot_spinbox = ttk.Spinbox(screenshot_config_frame, from_=1, to=20, width=8,
                                       textvariable=self.screenshot_count, font=self.fonts['body'])
        screenshot_spinbox.grid(row=0, column=1, padx=(10, 5))

        tk.Label(screenshot_section, text="(最高振幅的前N個時間點)",
               font=self.fonts['small'],
               fg=JAPANESE_COLORS['text_hint'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=2, column=0, columnspan=3, padx=15, pady=(0, 10))

    def create_analysis_settings_section(self):
        """創建分析設定區域 - 增強邊框效果並新增進階設定"""
        analysis_frame = ttk.LabelFrame(self.main_frame, text="🔧 分析設定",
                                       style='StrongCard.TLabelframe', padding="25")
        analysis_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 20), padx=10)
        analysis_frame.columnconfigure(0, weight=1)

        # 分析類型選擇區域 - 添加明顯邊框
        analysis_type_section = tk.Frame(analysis_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                       relief='solid', borderwidth=2,
                                       highlightbackground=JAPANESE_COLORS['border_strong'])
        analysis_type_section.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15), padx=5)

        analysis_label = tk.Label(analysis_type_section, text="🎯 分析類型:",
                                font=self.fonts['heading'],
                                fg=JAPANESE_COLORS['text_primary'],
                                bg=JAPANESE_COLORS['bg_secondary'])
        analysis_label.grid(row=0, column=0, sticky=tk.W, pady=(15, 10), padx=20)

        analysis_type_frame = tk.Frame(analysis_type_section, bg=JAPANESE_COLORS['bg_secondary'])
        analysis_type_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=20)

        # 高頻分析選項 - 添加邊框
        high_freq_frame = tk.Frame(analysis_type_frame, bg=JAPANESE_COLORS['bg_card'],
                                 relief='solid', borderwidth=1,
                                 highlightbackground=JAPANESE_COLORS['border'])
        high_freq_frame.grid(row=0, column=0, padx=(0, 10), pady=5)

        high_freq_radio = ttk.Radiobutton(high_freq_frame,
                                         text="高頻分析 (>1000 Hz)",
                                         variable=self.analysis_type,
                                         value="high_freq",
                                         command=self.on_analysis_type_change)
        high_freq_radio.pack(padx=15, pady=10)

        # 低頻分析選項 - 添加邊框
        low_freq_frame = tk.Frame(analysis_type_frame, bg=JAPANESE_COLORS['bg_card'],
                                relief='solid', borderwidth=1,
                                highlightbackground=JAPANESE_COLORS['border'])
        low_freq_frame.grid(row=0, column=1, padx=(0, 10), pady=5)

        low_freq_radio = ttk.Radiobutton(low_freq_frame,
                                        text="低頻分析 (<200 Hz)",
                                        variable=self.analysis_type,
                                        value="low_freq",
                                        command=self.on_analysis_type_change)
        low_freq_radio.pack(padx=15, pady=10)

        # 全頻分析選項 - 添加邊框
        full_freq_frame = tk.Frame(analysis_type_frame, bg=JAPANESE_COLORS['bg_card'],
                                 relief='solid', borderwidth=1,
                                 highlightbackground=JAPANESE_COLORS['border'])
        full_freq_frame.grid(row=0, column=2, padx=(0, 10), pady=5)

        full_freq_radio = ttk.Radiobutton(full_freq_frame,
                                         text="全頻分析 (20-20000 Hz)",
                                         variable=self.analysis_type,
                                         value="full_freq",
                                         command=self.on_analysis_type_change)
        full_freq_radio.pack(padx=15, pady=10)

        # 自訂頻率選項 - 添加邊框
        custom_freq_frame = tk.Frame(analysis_type_frame, bg=JAPANESE_COLORS['bg_card'],
                                    relief='solid', borderwidth=1,
                                    highlightbackground=JAPANESE_COLORS['border'])
        custom_freq_frame.grid(row=0, column=3, pady=5)

        custom_freq_radio = ttk.Radiobutton(custom_freq_frame,
                                           text="自訂頻率範圍",
                                           variable=self.analysis_type,
                                           value="custom",
                                           command=self.on_analysis_type_change)
        custom_freq_radio.pack(padx=15, pady=10)

        # 分析說明區域 - 添加邊框
        analysis_desc_section = tk.Frame(analysis_type_section, bg=JAPANESE_COLORS['bg_secondary'])
        analysis_desc_section.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 15), padx=20)

        descriptions = {
            "high_freq": "💡 適用於機械摩擦、軸承異音、電機高頻噪音等分析",
            "low_freq": "💡 適用於振動、共振、結構異音、低頻噪音等分析",
            "full_freq": "💡 完整頻譜分析，適用於綜合性音訊檢測",
            "custom": "💡 自訂特定頻率範圍，精確分析特定聲音"
        }

        self.analysis_desc_label = tk.Label(analysis_desc_section,
                                           text=descriptions["high_freq"],
                                           font=self.fonts['body'],
                                           fg=JAPANESE_COLORS['text_secondary'],
                                           bg=JAPANESE_COLORS['bg_secondary'])
        self.analysis_desc_label.grid(row=0, column=0, sticky=tk.W)

        # 🆕 自訂頻率範圍設定區域
        self.custom_freq_section = tk.Frame(analysis_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                           relief='solid', borderwidth=2,
                                           highlightbackground=JAPANESE_COLORS['border_strong'])
        self.custom_freq_section.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15), padx=5)
        self.custom_freq_section.grid_remove()  # 預設隱藏

        custom_title = tk.Label(self.custom_freq_section, text="🎚️ 自訂頻率範圍設定:",
                              font=self.fonts['heading'],
                              fg=JAPANESE_COLORS['text_primary'],
                              bg=JAPANESE_COLORS['bg_secondary'])
        custom_title.grid(row=0, column=0, columnspan=4, sticky=tk.W, pady=(15, 10), padx=20)

        # 頻率範圍輸入
        freq_input_frame = tk.Frame(self.custom_freq_section, bg=JAPANESE_COLORS['bg_secondary'])
        freq_input_frame.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 10), padx=20)

        tk.Label(freq_input_frame, text="起始頻率:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=0, sticky=tk.W, padx=(0, 10))

        self.custom_freq_start_entry = ttk.Spinbox(freq_input_frame, from_=20, to=20000, width=10,
                                                  textvariable=self.custom_freq_start, font=self.fonts['body'])
        self.custom_freq_start_entry.grid(row=0, column=1, padx=(0, 5))

        tk.Label(freq_input_frame, text="Hz",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=2, padx=(0, 30))

        tk.Label(freq_input_frame, text="結束頻率:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=3, sticky=tk.W, padx=(0, 10))

        self.custom_freq_end_entry = ttk.Spinbox(freq_input_frame, from_=20, to=20000, width=10,
                                                textvariable=self.custom_freq_end, font=self.fonts['body'])
        self.custom_freq_end_entry.grid(row=0, column=4, padx=(0, 5))

        tk.Label(freq_input_frame, text="Hz",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=5)

        # 常見頻率範圍快速選擇
        preset_frame = tk.Frame(self.custom_freq_section, bg=JAPANESE_COLORS['bg_secondary'])
        preset_frame.grid(row=2, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(5, 15), padx=20)

        tk.Label(preset_frame, text="💡 常見頻率預設:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=0, sticky=tk.W, pady=(0, 5))

        preset_buttons_frame = tk.Frame(preset_frame, bg=JAPANESE_COLORS['bg_secondary'])
        preset_buttons_frame.grid(row=1, column=0, sticky=tk.W)

        presets = [
            ("人聲範圍", 85, 255),
            ("音樂基頻", 200, 800),
            ("摩擦噪音", 1000, 5000),
            ("尖銳異音", 5000, 12000),
            ("超高頻", 12000, 20000)
        ]

        for i, (name, start, end) in enumerate(presets):
            btn = ttk.Button(preset_buttons_frame, text=name,
                           command=lambda s=start, e=end: self.set_custom_freq_preset(s, e),
                           style='Secondary.TButton')
            btn.grid(row=0, column=i, padx=(0, 5))

        # 🆕 時間範圍設定區域
        time_range_section = tk.Frame(analysis_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                     relief='solid', borderwidth=2,
                                     highlightbackground=JAPANESE_COLORS['border_strong'])
        time_range_section.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 15), padx=5)

        time_title = tk.Frame(time_range_section, bg=JAPANESE_COLORS['bg_secondary'])
        time_title.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(15, 10), padx=20)

        self.time_range_check = ttk.Checkbutton(time_title, text="⏰ 僅分析特定時間段",
                                               variable=self.use_time_range,
                                               command=self.toggle_time_range)
        self.time_range_check.grid(row=0, column=0, sticky=tk.W)

        # 時間範圍輸入
        self.time_input_frame = tk.Frame(time_range_section, bg=JAPANESE_COLORS['bg_secondary'])
        self.time_input_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15), padx=20)

        tk.Label(self.time_input_frame, text="開始時間:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=0, sticky=tk.W, padx=(0, 10))

        self.time_start_entry = ttk.Spinbox(self.time_input_frame, from_=0, to=9999, width=10,
                                           textvariable=self.time_start, font=self.fonts['body'],
                                           format="%.2f", increment=0.1, state='disabled')
        self.time_start_entry.grid(row=0, column=1, padx=(0, 5))

        tk.Label(self.time_input_frame, text="秒",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=2, padx=(0, 30))

        tk.Label(self.time_input_frame, text="結束時間:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=3, sticky=tk.W, padx=(0, 10))

        self.time_end_entry = ttk.Spinbox(self.time_input_frame, from_=0, to=9999, width=10,
                                         textvariable=self.time_end, font=self.fonts['body'],
                                         format="%.2f", increment=0.1, state='disabled')
        self.time_end_entry.grid(row=0, column=4, padx=(0, 5))

        tk.Label(self.time_input_frame, text="秒",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=5)

        tk.Label(self.time_input_frame, text="💡 只截取和分析此時間段內的音訊",
               font=self.fonts['small'],
               fg=JAPANESE_COLORS['text_hint'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=1, column=0, columnspan=6, sticky=tk.W, pady=(5, 0))

        # 🆕 振幅閾值設定區域
        threshold_section = tk.Frame(analysis_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                    relief='solid', borderwidth=2,
                                    highlightbackground=JAPANESE_COLORS['border_strong'])
        threshold_section.grid(row=3, column=0, sticky=(tk.W, tk.E), padx=5)

        threshold_title = tk.Frame(threshold_section, bg=JAPANESE_COLORS['bg_secondary'])
        threshold_title.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(15, 10), padx=20)

        self.threshold_check = ttk.Checkbutton(threshold_title, text="📊 設定振幅閾值 (只分析超過此強度的聲音)",
                                              variable=self.use_amplitude_threshold,
                                              command=self.toggle_amplitude_threshold)
        self.threshold_check.grid(row=0, column=0, sticky=tk.W)

        # 閾值輸入
        self.threshold_input_frame = tk.Frame(threshold_section, bg=JAPANESE_COLORS['bg_secondary'])
        self.threshold_input_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15), padx=20)

        tk.Label(self.threshold_input_frame, text="振幅閾值:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=0, sticky=tk.W, padx=(0, 10))

        self.threshold_entry = ttk.Spinbox(self.threshold_input_frame, from_=-80, to=0, width=10,
                                          textvariable=self.amplitude_threshold, font=self.fonts['body'],
                                          increment=5, state='disabled')
        self.threshold_entry.grid(row=0, column=1, padx=(0, 5))

        tk.Label(self.threshold_input_frame, text="dB",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=2, padx=(0, 20))

        # 閾值參考指南
        threshold_scale = tk.Scale(self.threshold_input_frame, from_=-80, to=0,
                                  orient=tk.HORIZONTAL, variable=self.amplitude_threshold,
                                  length=200, state='disabled', showvalue=0)
        threshold_scale.grid(row=0, column=3, padx=(0, 10))

        tk.Label(self.threshold_input_frame, text="💡 -60dB: 微弱聲音 | -40dB: 中等聲音 | -20dB: 強烈聲音",
               font=self.fonts['small'],
               fg=JAPANESE_COLORS['text_hint'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=1, column=0, columnspan=4, sticky=tk.W, pady=(5, 0))

        self.threshold_scale = threshold_scale  # 保存引用

        # 🆕 多頻帶同步分析設定
        multi_band_section = tk.Frame(analysis_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                     relief='solid', borderwidth=2,
                                     highlightbackground=JAPANESE_COLORS['border_strong'])
        multi_band_section.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(15, 15), padx=5)

        multi_band_title = tk.Frame(multi_band_section, bg=JAPANESE_COLORS['bg_secondary'])
        multi_band_title.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(15, 10), padx=20)

        self.multi_band_check = ttk.Checkbutton(multi_band_title,
                                               text="🎼 多頻帶同步分析 (同時監控多個聲音來源)",
                                               variable=self.use_multi_band,
                                               command=self.toggle_multi_band)
        self.multi_band_check.grid(row=0, column=0, sticky=tk.W)

        # 多頻帶設定區域
        self.multi_band_config = tk.Frame(multi_band_section, bg=JAPANESE_COLORS['bg_secondary'])
        self.multi_band_config.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15), padx=20)

        tk.Label(self.multi_band_config, text="💡 定義多個頻率範圍以分離不同聲音來源（例如：低頻馬達 + 高頻摩擦）",
               font=self.fonts['small'],
               fg=JAPANESE_COLORS['text_hint'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=0, columnspan=4, sticky=tk.W, pady=(0, 10))

        # 頻帶列表顯示區域
        self.band_list_frame = tk.Frame(self.multi_band_config, bg=JAPANESE_COLORS['bg_card'],
                                       relief='solid', borderwidth=1)
        self.band_list_frame.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(0, 10))

        self.band_list_label = tk.Label(self.band_list_frame, text="尚未新增頻帶",
                                       font=self.fonts['body'],
                                       fg=JAPANESE_COLORS['text_hint'],
                                       bg=JAPANESE_COLORS['bg_card'])
        self.band_list_label.pack(padx=15, pady=10)

        # 新增頻帶控制
        add_band_frame = tk.Frame(self.multi_band_config, bg=JAPANESE_COLORS['bg_secondary'])
        add_band_frame.grid(row=2, column=0, columnspan=4, sticky=tk.W, pady=(0, 5))

        tk.Label(add_band_frame, text="新增頻帶:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=0, padx=(0, 10))

        self.new_band_start = tk.IntVar(value=100)
        self.new_band_end = tk.IntVar(value=500)
        self.new_band_name = tk.StringVar(value="")

        tk.Label(add_band_frame, text="名稱:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=1)

        ttk.Entry(add_band_frame, textvariable=self.new_band_name, width=12,
                 font=self.fonts['body']).grid(row=0, column=2, padx=(5, 15))

        tk.Label(add_band_frame, text="範圍:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=3)

        ttk.Spinbox(add_band_frame, from_=20, to=20000, width=8,
                   textvariable=self.new_band_start, font=self.fonts['body']).grid(row=0, column=4, padx=(5, 2))

        tk.Label(add_band_frame, text="-",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=5, padx=2)

        ttk.Spinbox(add_band_frame, from_=20, to=20000, width=8,
                   textvariable=self.new_band_end, font=self.fonts['body']).grid(row=0, column=6, padx=(2, 5))

        tk.Label(add_band_frame, text="Hz",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=7, padx=(0, 15))

        ttk.Button(add_band_frame, text="➕ 新增頻帶",
                  command=self.add_frequency_band, style='Secondary.TButton').grid(row=0, column=8)

        ttk.Button(add_band_frame, text="🗑️ 清空全部",
                  command=self.clear_frequency_bands, style='Secondary.TButton').grid(row=0, column=9, padx=(5, 0))

        # 快速預設組合
        preset_combo_frame = tk.Frame(self.multi_band_config, bg=JAPANESE_COLORS['bg_secondary'])
        preset_combo_frame.grid(row=3, column=0, columnspan=4, sticky=tk.W, pady=(5, 0))

        tk.Label(preset_combo_frame, text="快速預設:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=0, padx=(0, 10))

        ttk.Button(preset_combo_frame, text="機械三頻帶",
                  command=lambda: self.load_preset_bands("mechanical"),
                  style='Secondary.TButton').grid(row=0, column=1, padx=(0, 5))

        ttk.Button(preset_combo_frame, text="音訊分層",
                  command=lambda: self.load_preset_bands("audio"),
                  style='Secondary.TButton').grid(row=0, column=2, padx=(0, 5))

        # 初始狀態為停用
        for widget in self.multi_band_config.winfo_children():
            self.set_widget_state(widget, 'disabled')

        # 🆕 音訊節錄設定
        audio_extract_section = tk.Frame(analysis_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                        relief='solid', borderwidth=2,
                                        highlightbackground=JAPANESE_COLORS['border_strong'])
        audio_extract_section.grid(row=5, column=0, sticky=(tk.W, tk.E), padx=5)

        audio_title = tk.Frame(audio_extract_section, bg=JAPANESE_COLORS['bg_secondary'])
        audio_title.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(15, 10), padx=20)

        self.audio_extract_check = ttk.Checkbutton(audio_title,
                                                   text="🎵 自動節錄音訊片段 (擷取峰值時刻的聲音)",
                                                   variable=self.enable_audio_extract,
                                                   command=self.toggle_audio_extract)
        self.audio_extract_check.grid(row=0, column=0, sticky=tk.W)

        # 音訊節錄參數
        self.audio_extract_config = tk.Frame(audio_extract_section, bg=JAPANESE_COLORS['bg_secondary'])
        self.audio_extract_config.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15), padx=20)

        tk.Label(self.audio_extract_config, text="節錄範圍:",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=0, sticky=tk.W, padx=(0, 10))

        tk.Label(self.audio_extract_config, text="峰值前",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=1, padx=(0, 5))

        self.audio_before_entry = ttk.Spinbox(self.audio_extract_config, from_=0.1, to=10, width=6,
                                             textvariable=self.audio_duration_before,
                                             font=self.fonts['body'],
                                             format="%.1f", increment=0.5)
        self.audio_before_entry.grid(row=0, column=2, padx=(0, 5))

        tk.Label(self.audio_extract_config, text="秒  +  峰值後",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=3, padx=(0, 5))

        self.audio_after_entry = ttk.Spinbox(self.audio_extract_config, from_=0.1, to=10, width=6,
                                            textvariable=self.audio_duration_after,
                                            font=self.fonts['body'],
                                            format="%.1f", increment=0.5)
        self.audio_after_entry.grid(row=0, column=4, padx=(0, 5))

        tk.Label(self.audio_extract_config, text="秒",
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=0, column=5)

        tk.Label(self.audio_extract_config,
               text="💡 將自動儲存為 WAV 檔案，並可在截圖預覽區域播放試聽",
               font=self.fonts['small'],
               fg=JAPANESE_COLORS['text_hint'],
               bg=JAPANESE_COLORS['bg_secondary']).grid(row=1, column=0, columnspan=6, sticky=tk.W, pady=(5, 0))

        # 綁定分析類型變更事件
        self.analysis_type.trace('w', lambda *args: self.update_analysis_description())

    def set_widget_state(self, widget, state):
        """遞迴設定widget及其子widget的狀態"""
        try:
            widget.configure(state=state)
        except:
            pass
        for child in widget.winfo_children():
            self.set_widget_state(child, state)

    def toggle_multi_band(self):
        """切換多頻帶分析設定"""
        if self.use_multi_band.get():
            for widget in self.multi_band_config.winfo_children():
                self.set_widget_state(widget, 'normal')
        else:
            for widget in self.multi_band_config.winfo_children():
                self.set_widget_state(widget, 'disabled')

    def toggle_audio_extract(self):
        """切換音訊節錄設定"""
        if self.enable_audio_extract.get():
            self.audio_before_entry.config(state='normal')
            self.audio_after_entry.config(state='normal')
        else:
            self.audio_before_entry.config(state='disabled')
            self.audio_after_entry.config(state='disabled')

    def add_frequency_band(self):
        """新增頻率範圍"""
        start = self.new_band_start.get()
        end = self.new_band_end.get()
        name = self.new_band_name.get().strip()

        if start >= end:
            messagebox.showwarning("警告", "起始頻率必須小於結束頻率")
            return

        if not name:
            name = f"頻帶{len(self.frequency_bands) + 1}"

        # 分配顏色
        colors = ['#E57373', '#81C784', '#64B5F6', '#FFB74D', '#BA68C8', '#4DB6AC']
        color = colors[len(self.frequency_bands) % len(colors)]

        band = {
            'name': name,
            'start': start,
            'end': end,
            'color': color
        }

        self.frequency_bands.append(band)
        self.update_band_list()

        # 重置輸入
        self.new_band_name.set("")

    def clear_frequency_bands(self):
        """清空所有頻率範圍"""
        self.frequency_bands = []
        self.update_band_list()

    def update_band_list(self):
        """更新頻帶列表顯示"""
        # 清空現有顯示
        for widget in self.band_list_frame.winfo_children():
            widget.destroy()

        if not self.frequency_bands:
            self.band_list_label = tk.Label(self.band_list_frame, text="尚未新增頻帶",
                                           font=self.fonts['body'],
                                           fg=JAPANESE_COLORS['text_hint'],
                                           bg=JAPANESE_COLORS['bg_card'])
            self.band_list_label.pack(padx=15, pady=10)
        else:
            for i, band in enumerate(self.frequency_bands):
                band_frame = tk.Frame(self.band_list_frame, bg=JAPANESE_COLORS['bg_card'])
                band_frame.pack(fill=tk.X, padx=10, pady=5)

                # 顏色標記
                color_label = tk.Label(band_frame, text="█", font=('Arial', 14),
                                     fg=band['color'], bg=JAPANESE_COLORS['bg_card'])
                color_label.pack(side=tk.LEFT, padx=(5, 10))

                # 頻帶資訊
                info_text = f"{band['name']}: {band['start']}-{band['end']} Hz"
                info_label = tk.Label(band_frame, text=info_text,
                                    font=self.fonts['body'],
                                    fg=JAPANESE_COLORS['text_primary'],
                                    bg=JAPANESE_COLORS['bg_card'])
                info_label.pack(side=tk.LEFT)

                # 刪除按鈕
                del_btn = ttk.Button(band_frame, text="✕", width=3,
                                   command=lambda idx=i: self.remove_frequency_band(idx),
                                   style='Secondary.TButton')
                del_btn.pack(side=tk.RIGHT, padx=5)

    def remove_frequency_band(self, index):
        """移除指定的頻率範圍"""
        if 0 <= index < len(self.frequency_bands):
            self.frequency_bands.pop(index)
            self.update_band_list()

    def load_preset_bands(self, preset_type):
        """載入預設的頻帶組合"""
        self.frequency_bands = []

        if preset_type == "mechanical":
            # 機械分析三頻帶
            self.frequency_bands = [
                {'name': '低頻振動', 'start': 20, 'end': 200, 'color': '#E57373'},
                {'name': '中頻運轉', 'start': 200, 'end': 2000, 'color': '#81C784'},
                {'name': '高頻摩擦', 'start': 2000, 'end': 10000, 'color': '#64B5F6'}
            ]
        elif preset_type == "audio":
            # 音訊分層
            self.frequency_bands = [
                {'name': '低音', 'start': 20, 'end': 250, 'color': '#E57373'},
                {'name': '中音', 'start': 250, 'end': 2000, 'color': '#81C784'},
                {'name': '高音', 'start': 2000, 'end': 6000, 'color': '#64B5F6'},
                {'name': '超高音', 'start': 6000, 'end': 20000, 'color': '#FFB74D'}
            ]

        self.update_band_list()

    def on_analysis_type_change(self):
        """分析類型改變時的處理"""
        if self.analysis_type.get() == "custom":
            self.custom_freq_section.grid()
        else:
            self.custom_freq_section.grid_remove()
        self.update_analysis_description()

    def set_custom_freq_preset(self, start, end):
        """設定預設頻率範圍"""
        self.custom_freq_start.set(start)
        self.custom_freq_end.set(end)

    def toggle_time_range(self):
        """切換時間範圍設定"""
        if self.use_time_range.get():
            self.time_start_entry.config(state='normal')
            self.time_end_entry.config(state='normal')
            # 如果已經選擇影片，自動設定結束時間為影片長度
            if self.audio_duration > 0:
                self.time_end.set(self.audio_duration)
        else:
            self.time_start_entry.config(state='disabled')
            self.time_end_entry.config(state='disabled')

    def toggle_amplitude_threshold(self):
        """切換振幅閾值設定"""
        if self.use_amplitude_threshold.get():
            self.threshold_entry.config(state='normal')
            self.threshold_scale.config(state='normal')
        else:
            self.threshold_entry.config(state='disabled')
            self.threshold_scale.config(state='disabled')

    def update_analysis_description(self):
        """更新分析類型說明"""
        descriptions = {
            "high_freq": "💡 適用於機械摩擦、軸承異音、電機高頻噪音等分析",
            "low_freq": "💡 適用於振動、共振、結構異音、低頻噪音等分析",
            "full_freq": "💡 完整頻譜分析，適用於綜合性音訊檢測",
            "custom": "💡 自訂特定頻率範圍，精確分析特定聲音"
        }

        current_type = self.analysis_type.get()
        self.analysis_desc_label.config(text=descriptions.get(current_type, ""))

    def create_control_section(self):
        """創建控制區域 - 增強視覺效果"""
        control_frame = ttk.LabelFrame(self.main_frame, text="⚡ 步驟 2: 開始分析",
                                      style='StrongCard.TLabelframe', padding="25")
        control_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 20), padx=10)
        control_frame.columnconfigure(1, weight=1)

        # 按鈕容器
        button_container = tk.Frame(control_frame, bg=JAPANESE_COLORS['bg_card'])
        button_container.grid(row=0, column=0, padx=(0, 20), pady=10)

        # 分析按鈕區域 - 添加強調邊框
        analyze_section = tk.Frame(button_container, bg=JAPANESE_COLORS['accent_primary'],
                                relief='solid', borderwidth=3,
                                highlightbackground=JAPANESE_COLORS['border_strong'])
        analyze_section.pack(pady=(0, 10))

        self.process_button = ttk.Button(analyze_section, text="🎬 開始頻譜分析與截圖",
                                        command=self.start_analysis, state="disabled",
                                        style='Primary.TButton')
        self.process_button.pack(padx=20, pady=15)

        # 🆕 導出按鈕區域
        export_section = tk.Frame(button_container, bg=JAPANESE_COLORS['bg_secondary'],
                                relief='solid', borderwidth=2,
                                highlightbackground=JAPANESE_COLORS['border'])
        export_section.pack()

        self.export_button = ttk.Button(export_section, text="📊 導出 Excel 分析報告",
                                       command=self.export_to_excel, state="disabled",
                                       style='Secondary.TButton')
        self.export_button.pack(padx=20, pady=10)

        tk.Label(export_section, text="💡 將分析結果匯出成 Excel 檔案",
               font=self.fonts['small'],
               fg=JAPANESE_COLORS['text_hint'],
               bg=JAPANESE_COLORS['bg_secondary']).pack(padx=20, pady=(0, 10))

        # 進度區域 - 添加邊框
        progress_section = tk.Frame(control_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                  relief='solid', borderwidth=1,
                                  highlightbackground=JAPANESE_COLORS['border'])
        progress_section.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=10)
        progress_section.columnconfigure(0, weight=1)

        progress_title = tk.Label(progress_section, text="📊 處理進度",
                                font=self.fonts['heading'],
                                fg=JAPANESE_COLORS['text_primary'],
                                bg=JAPANESE_COLORS['bg_secondary'])
        progress_title.grid(row=0, column=0, pady=(10, 5))

        self.progress = ttk.Progressbar(progress_section, mode='determinate',
                                       style='Japanese.Horizontal.TProgressbar')
        self.progress.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=20, pady=(0, 5))

        self.status_label = tk.Label(progress_section, text="等待開始分析...",
                                    font=self.fonts['body'],
                                    fg=JAPANESE_COLORS['text_primary'],
                                    bg=JAPANESE_COLORS['bg_secondary'])
        self.status_label.grid(row=2, column=0, padx=20, pady=(0, 10))

    def create_preview_section(self):
        """創建預覽區域 - 增強邊框"""
        preview_frame = ttk.LabelFrame(self.main_frame, text="🖼️ 步驟 3: 截圖預覽",
                                      style='StrongCard.TLabelframe', padding="20")
        preview_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(0, 20), padx=10)
        preview_frame.columnconfigure(0, weight=1)

        # 創建預覽容器 - 添加明顯邊框
        preview_container = tk.Frame(preview_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                   relief='solid', borderwidth=2,
                                   highlightbackground=JAPANESE_COLORS['border_strong'])
        preview_container.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=10)
        preview_container.columnconfigure(0, weight=1)

        # 創建滾動框架用於顯示截圖
        preview_canvas = tk.Canvas(preview_container, height=160, bg=JAPANESE_COLORS['bg_secondary'],
                                  highlightthickness=0,
                                  relief='sunken', borderwidth=1)
        preview_scrollbar = ttk.Scrollbar(preview_container, orient="horizontal",
                                         command=preview_canvas.xview)

        self.preview_scrollable_frame = tk.Frame(preview_canvas, bg=JAPANESE_COLORS['bg_secondary'])

        self.preview_scrollable_frame.bind(
            "<Configure>",
            lambda e: preview_canvas.configure(scrollregion=preview_canvas.bbox("all"))
        )

        preview_canvas.create_window((0, 0), window=self.preview_scrollable_frame, anchor="nw")
        preview_canvas.configure(xscrollcommand=preview_scrollbar.set)

        preview_canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=(10, 5))
        preview_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=10, pady=(0, 10))

    def create_results_section(self):
        """創建結果區域 - 增強邊框效果"""
        result_frame = ttk.LabelFrame(self.main_frame, text="📊 步驟 4: 分析結果",
                                     style='StrongCard.TLabelframe', padding="20")
        result_frame.grid(row=5, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=(0, 10))
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)

        # 創建結果容器 - 添加邊框
        results_container = tk.Frame(result_frame, bg=JAPANESE_COLORS['bg_card'],
                                   relief='solid', borderwidth=2,
                                   highlightbackground=JAPANESE_COLORS['border_strong'])
        results_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        results_container.columnconfigure(0, weight=1)
        results_container.rowconfigure(0, weight=1)

        # 創建筆記本（標籤頁）- 增強邊框
        self.notebook = ttk.Notebook(results_container, style='Japanese.TNotebook')
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=10)

        # 📊 視覺化總覽標籤頁（新增 - 最簡單易懂）
        self.summary_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.summary_frame, text="📊 視覺化總覽")

        # 頻譜圖標籤頁
        self.spectrum_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.spectrum_frame, text="📈 整體頻譜圖")

        # 時頻分析標籤頁
        self.timefreq_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.timefreq_frame, text="🌊 時頻分析圖 (進階)")

        # 分析結果標籤頁
        self.analysis_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.analysis_frame, text="📋 分析報告")

        # 創建matplotlib圖表
        self.create_plots()

        # 創建視覺化總覽
        self.create_visual_summary()

        # 創建分析結果顯示區域
        self.create_analysis_display()

    def create_visual_summary(self):
        """創建視覺化總覽 - 簡單易懂的呈現方式"""
        # 創建可滾動的畫布
        summary_canvas = tk.Canvas(self.summary_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                  highlightthickness=0)
        summary_scrollbar = ttk.Scrollbar(self.summary_frame, orient="vertical",
                                         command=summary_canvas.yview)

        self.summary_content = tk.Frame(summary_canvas, bg=JAPANESE_COLORS['bg_secondary'])

        self.summary_content.bind(
            "<Configure>",
            lambda e: summary_canvas.configure(scrollregion=summary_canvas.bbox("all"))
        )

        summary_canvas.create_window((0, 0), window=self.summary_content, anchor="nw")
        summary_canvas.configure(yscrollcommand=summary_scrollbar.set)

        summary_canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        summary_scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)

        # 綁定滾輪
        def _on_mousewheel(event):
            summary_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        summary_canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def update_visual_summary(self):
        """更新視覺化總覽內容"""
        # 清空現有內容
        for widget in self.summary_content.winfo_children():
            widget.destroy()

        # 標題區域
        title_frame = tk.Frame(self.summary_content, bg=JAPANESE_COLORS['bg_card'],
                             relief='solid', borderwidth=2,
                             highlightbackground=JAPANESE_COLORS['border_strong'])
        title_frame.pack(fill=tk.X, padx=20, pady=(10, 20))

        title_label = tk.Label(title_frame,
                              text="📊 分析結果視覺化總覽 - 簡單易懂版",
                              font=self.fonts['heading'],
                              fg=JAPANESE_COLORS['text_primary'],
                              bg=JAPANESE_COLORS['bg_card'])
        title_label.pack(pady=15)

        subtitle_label = tk.Label(title_frame,
                                 text="💡 這個頁面用簡單的方式呈現分析結果，適合初學者快速理解",
                                 font=self.fonts['body'],
                                 fg=JAPANESE_COLORS['text_secondary'],
                                 bg=JAPANESE_COLORS['bg_card'])
        subtitle_label.pack(pady=(0, 15))

        # 🎯 關鍵發現區域
        key_findings = tk.Frame(self.summary_content, bg=JAPANESE_COLORS['bg_card'],
                              relief='solid', borderwidth=2,
                              highlightbackground=JAPANESE_COLORS['border_strong'])
        key_findings.pack(fill=tk.X, padx=20, pady=(0, 20))

        tk.Label(key_findings, text="🎯 關鍵發現",
               font=self.fonts['heading'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_card']).pack(pady=(15, 10))

        # 統計卡片容器
        stats_container = tk.Frame(key_findings, bg=JAPANESE_COLORS['bg_card'])
        stats_container.pack(fill=tk.X, padx=20, pady=(0, 15))

        # 創建統計卡片
        if self.global_max_peak:
            self.create_stat_card(stats_container,
                                "⭐ 最強聲音出現時間",
                                f"{self.global_max_peak['time']:.2f} 秒",
                                f"強度: {self.global_max_peak['max_magnitude']:.1f} dB",
                                JAPANESE_COLORS['error'], 0)

            self.create_stat_card(stats_container,
                                "🎵 最強聲音的頻率",
                                f"{self.global_max_peak['max_frequency']:.0f} Hz",
                                self.get_frequency_description(self.global_max_peak['max_frequency']),
                                JAPANESE_COLORS['accent_primary'], 1)

        if self.captured_screenshots:
            self.create_stat_card(stats_container,
                                "📸 重點時刻截圖",
                                f"{len(self.captured_screenshots)} 張",
                                "已標記異常聲音發生時的畫面",
                                JAPANESE_COLORS['accent_secondary'], 2)

        total_peaks = len(self.time_peaks)
        self.create_stat_card(stats_container,
                            "🔍 檢測到的異常時刻",
                            f"{total_peaks} 個",
                            "影片中有異常聲音的時間點",
                            JAPANESE_COLORS['warning'], 3)

        # 📈 聲音強度時間軸
        timeline = tk.Frame(self.summary_content, bg=JAPANESE_COLORS['bg_card'],
                          relief='solid', borderwidth=2,
                          highlightbackground=JAPANESE_COLORS['border_strong'])
        timeline.pack(fill=tk.X, padx=20, pady=(0, 20))

        tk.Label(timeline, text="📈 聲音強度時間軸",
               font=self.fonts['heading'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_card']).pack(pady=(15, 5))

        tk.Label(timeline, text="💡 高度代表聲音強度，顏色越紅表示越響亮",
               font=self.fonts['small'],
               fg=JAPANESE_COLORS['text_hint'],
               bg=JAPANESE_COLORS['bg_card']).pack(pady=(0, 10))

        self.create_simple_timeline(timeline)

        # 🎼 多頻帶對比（如果有）
        if self.use_multi_band.get() and self.multi_band_peaks:
            multi_band = tk.Frame(self.summary_content, bg=JAPANESE_COLORS['bg_card'],
                                relief='solid', borderwidth=2,
                                highlightbackground=JAPANESE_COLORS['border_strong'])
            multi_band.pack(fill=tk.X, padx=20, pady=(0, 20))

            tk.Label(multi_band, text="🎼 各頻帶聲音強度比較",
                   font=self.fonts['heading'],
                   fg=JAPANESE_COLORS['text_primary'],
                   bg=JAPANESE_COLORS['bg_card']).pack(pady=(15, 5))

            tk.Label(multi_band, text="💡 比較不同類型聲音的強度分布",
                   font=self.fonts['small'],
                   fg=JAPANESE_COLORS['text_hint'],
                   bg=JAPANESE_COLORS['bg_card']).pack(pady=(0, 10))

            self.create_band_comparison(multi_band)

        # 📋 簡易判讀說明
        guide = tk.Frame(self.summary_content, bg=JAPANESE_COLORS['bg_card'],
                       relief='solid', borderwidth=2,
                       highlightbackground=JAPANESE_COLORS['border_strong'])
        guide.pack(fill=tk.X, padx=20, pady=(0, 20))

        tk.Label(guide, text="📋 如何判讀結果",
               font=self.fonts['heading'],
               fg=JAPANESE_COLORS['text_primary'],
               bg=JAPANESE_COLORS['bg_card']).pack(pady=(15, 10))

        guide_texts = [
            "🔴 紅色時刻 = 聲音最響亮的時候，需要特別注意",
            "🟡 黃色時刻 = 中等強度，可能有異常",
            "🟢 綠色時刻 = 聲音正常，無需擔心",
            "📸 有截圖的時刻 = 程式自動捕捉的重點畫面",
            "🎵 可以播放音訊 = 確認聲音是否為您要找的異音"
        ]

        for text in guide_texts:
            label = tk.Label(guide, text=text,
                           font=self.fonts['body'],
                           fg=JAPANESE_COLORS['text_primary'],
                           bg=JAPANESE_COLORS['bg_card'],
                           anchor='w')
            label.pack(fill=tk.X, padx=30, pady=3)

        tk.Label(guide, text=" ", bg=JAPANESE_COLORS['bg_card']).pack(pady=10)

    def create_stat_card(self, parent, title, value, description, color, column):
        """創建統計卡片"""
        card = tk.Frame(parent, bg=color, relief='solid', borderwidth=0)
        card.grid(row=0, column=column, padx=10, pady=10, sticky=(tk.N, tk.S, tk.E, tk.W))
        parent.columnconfigure(column, weight=1)

        # 內容框架
        content = tk.Frame(card, bg='white')
        content.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)

        tk.Label(content, text=title,
               font=self.fonts['body'],
               fg=JAPANESE_COLORS['text_secondary'],
               bg='white').pack(pady=(15, 5))

        tk.Label(content, text=value,
               font=(self.font_family, 20, 'bold'),
               fg=color,
               bg='white').pack(pady=5)

        tk.Label(content, text=description,
               font=self.fonts['small'],
               fg=JAPANESE_COLORS['text_hint'],
               bg='white',
               wraplength=200).pack(pady=(5, 15))

    def create_simple_timeline(self, parent):
        """創建簡單的時間軸視覺化"""
        if not self.time_peaks:
            tk.Label(parent, text="尚無分析資料",
                   font=self.fonts['body'],
                   fg=JAPANESE_COLORS['text_hint'],
                   bg=JAPANESE_COLORS['bg_card']).pack(pady=20)
            return

        timeline_canvas = tk.Canvas(parent, height=200, bg='white',
                                   highlightthickness=0)
        timeline_canvas.pack(fill=tk.X, padx=20, pady=(0, 20))

        # 計算時間範圍
        times = [p['time'] for p in self.time_peaks]
        magnitudes = [p['max_magnitude'] for p in self.time_peaks]

        if not times:
            return

        min_time = min(times)
        max_time = max(times)
        min_mag = min(magnitudes)
        max_mag = max(magnitudes)

        # 繪製時間軸
        width = timeline_canvas.winfo_reqwidth()
        if width <= 1:
            width = 800

        height = 180
        margin = 40

        # 繪製背景網格
        for i in range(5):
            y = margin + i * (height - 2 * margin) / 4
            timeline_canvas.create_line(margin, y, width - margin, y,
                                      fill=JAPANESE_COLORS['divider'], width=1)

        # 繪製時間軸線
        timeline_canvas.create_line(margin, height - margin,
                                   width - margin, height - margin,
                                   fill=JAPANESE_COLORS['text_secondary'], width=2)

        # 繪製數據點
        screenshot_times = {s['time'] for s in self.captured_screenshots}

        for i, (time, mag) in enumerate(zip(times, magnitudes)):
            # 計算位置
            if max_time > min_time:
                x = margin + (time - min_time) / (max_time - min_time) * (width - 2 * margin)
            else:
                x = width / 2

            # 計算高度（振幅）
            if max_mag > min_mag:
                normalized_mag = (mag - min_mag) / (max_mag - min_mag)
            else:
                normalized_mag = 0.5

            bar_height = normalized_mag * (height - 2 * margin)
            y_top = height - margin - bar_height

            # 根據振幅選擇顏色
            if mag > -30:
                color = '#E57373'  # 紅色 - 很響
            elif mag > -40:
                color = '#FFB74D'  # 橙色 - 中等
            else:
                color = '#81C784'  # 綠色 - 正常

            # 繪製柱狀
            bar_width = max(2, (width - 2 * margin) / len(times) * 0.8)
            timeline_canvas.create_rectangle(x - bar_width/2, y_top,
                                           x + bar_width/2, height - margin,
                                           fill=color, outline='')

            # 如果有截圖，標記星號
            if time in screenshot_times:
                timeline_canvas.create_text(x, y_top - 10, text="⭐",
                                          font=('Arial', 12))

        # 添加時間標籤
        for i in range(5):
            if max_time > min_time:
                time_val = min_time + i * (max_time - min_time) / 4
            else:
                time_val = min_time
            x = margin + i * (width - 2 * margin) / 4
            timeline_canvas.create_text(x, height - margin + 15,
                                       text=f"{time_val:.1f}s",
                                       font=self.fonts['small'],
                                       fill=JAPANESE_COLORS['text_secondary'])

    def create_band_comparison(self, parent):
        """創建頻帶比較圖"""
        if not self.multi_band_peaks:
            return

        comparison_frame = tk.Frame(parent, bg='white')
        comparison_frame.pack(fill=tk.X, padx=20, pady=(0, 20))

        # 為每個頻帶創建比較條
        for band_name, band_data in self.multi_band_peaks.items():
            band_frame = tk.Frame(comparison_frame, bg='white')
            band_frame.pack(fill=tk.X, pady=8)

            # 頻帶名稱
            name_label = tk.Label(band_frame, text=band_name,
                                font=self.fonts['body'],
                                fg=JAPANESE_COLORS['text_primary'],
                                bg='white', width=15, anchor='w')
            name_label.pack(side=tk.LEFT, padx=(10, 10))

            # 強度條
            max_peak = band_data.get('max_peak')
            if max_peak:
                magnitude = max_peak['max_magnitude']
                # 轉換為 0-100 的比例（假設 -80 到 0 dB）
                percentage = max(0, min(100, (magnitude + 80) / 80 * 100))

                bar_container = tk.Frame(band_frame, bg=JAPANESE_COLORS['divider'],
                                       height=30, width=400)
                bar_container.pack(side=tk.LEFT, padx=(0, 10))
                bar_container.pack_propagate(False)

                bar = tk.Frame(bar_container, bg=band_data['color'], height=30)
                bar.place(x=0, y=0, relwidth=percentage/100, relheight=1)

                # 數值標籤
                value_label = tk.Label(band_frame,
                                     text=f"{magnitude:.1f} dB ({percentage:.0f}%)",
                                     font=self.fonts['body'],
                                     fg=JAPANESE_COLORS['text_primary'],
                                     bg='white')
                value_label.pack(side=tk.LEFT)

    def get_frequency_description(self, freq):
        """根據頻率返回描述"""
        if freq < 200:
            return "低頻聲音（振動、低鳴）"
        elif freq < 800:
            return "中低頻（人聲、運轉聲）"
        elif freq < 2000:
            return "中頻（對話、音樂）"
        elif freq < 5000:
            return "中高頻（摩擦聲）"
        elif freq < 10000:
            return "高頻（尖銳聲、碰撞）"
        else:
            return "超高頻（極尖銳聲響）"

    def create_plots(self):
        # 創建整體頻譜圖容器 - 添加邊框
        spectrum_container = tk.Frame(self.spectrum_frame, bg=JAPANESE_COLORS['bg_card'],
                                    relief='solid', borderwidth=1,
                                    highlightbackground=JAPANESE_COLORS['border'])
        spectrum_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 創建整體頻譜圖
        self.fig1 = Figure(figsize=(12, 6), dpi=100, facecolor=JAPANESE_COLORS['bg_card'])
        self.ax1 = self.fig1.add_subplot(111)
        self.ax1.set_title("音訊頻譜分析 (dB)", fontsize=14, color=JAPANESE_COLORS['text_primary'], pad=20)
        self.ax1.set_xlabel("頻率 (Hz)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
        self.ax1.set_ylabel("振幅 (dB)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
        self.ax1.grid(True, which="both", linestyle='--', linewidth=0.5, color=JAPANESE_COLORS['divider'])
        self.ax1.set_xlim(0, 20000)
        self.ax1.set_facecolor(JAPANESE_COLORS['bg_secondary'])

        # 創建畫布1
        self.canvas1 = FigureCanvasTkAgg(self.fig1, spectrum_container)
        self.canvas1.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 創建時頻分析圖容器 - 添加邊框
        timefreq_container = tk.Frame(self.timefreq_frame, bg=JAPANESE_COLORS['bg_card'],
                                    relief='solid', borderwidth=1,
                                    highlightbackground=JAPANESE_COLORS['border'])
        timefreq_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 創建時頻分析圖
        self.fig2 = Figure(figsize=(12, 8), dpi=100, facecolor=JAPANESE_COLORS['bg_card'])

        # 時頻圖（頻譜圖）
        self.ax2 = self.fig2.add_subplot(211)
        self.ax2.set_title("時頻分析圖 (頻譜圖)", fontsize=14, color=JAPANESE_COLORS['text_primary'], pad=15)
        self.ax2.set_xlabel("時間 (秒)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
        self.ax2.set_ylabel("頻率 (Hz)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
        self.ax2.set_facecolor(JAPANESE_COLORS['bg_secondary'])

        # 時間-峰值振幅圖
        self.ax3 = self.fig2.add_subplot(212)
        self.ax3.set_title("各時間點最大振幅變化", fontsize=14, color=JAPANESE_COLORS['text_primary'], pad=15)
        self.ax3.set_xlabel("時間 (秒)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
        self.ax3.set_ylabel("最大振幅 (dB)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
        self.ax3.grid(True, linestyle='--', linewidth=0.5, color=JAPANESE_COLORS['divider'])
        self.ax3.set_facecolor(JAPANESE_COLORS['bg_secondary'])

        self.fig2.tight_layout(pad=3.0)

        # 創建畫布2
        self.canvas2 = FigureCanvasTkAgg(self.fig2, timefreq_container)
        self.canvas2.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def create_analysis_display(self):
        """創建分析結果顯示區域 - 增強邊框"""
        # 分析結果容器 - 添加邊框
        analysis_container = tk.Frame(self.analysis_frame, bg=JAPANESE_COLORS['bg_card'],
                                    relief='solid', borderwidth=1,
                                    highlightbackground=JAPANESE_COLORS['border'])
        analysis_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        text_frame = tk.Frame(analysis_container, bg=JAPANESE_COLORS['bg_card'])
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 滾動條
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 文本框 - 增強邊框
        self.result_text = tk.Text(text_frame, wrap=tk.WORD,
                                  yscrollcommand=scrollbar.set,
                                  font=self.fonts['body'],
                                  bg=JAPANESE_COLORS['bg_secondary'],
                                  fg=JAPANESE_COLORS['text_primary'],
                                  selectbackground=JAPANESE_COLORS['accent_primary'],
                                  selectforeground='white',
                                  borderwidth=2,
                                  relief='solid',
                                  highlightbackground=JAPANESE_COLORS['border'],
                                  padx=15, pady=15)
        self.result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.result_text.yview)

    # 保留原有的其他方法（分析邏輯部分不需要修改邊框相關代碼）
    def select_file(self):
        """選擇影片檔案"""
        file_types = [
            ("影片檔案", "*.mp4 *.avi *.mov *.mkv *.wmv *.flv *.webm *.m4v *.3gp *.mpg *.mpeg"),
            ("常用格式", "*.mp4 *.avi *.mov *.mkv"),
            ("所有檔案", "*.*")
        ]

        filename = filedialog.askopenfilename(
            title="選擇影片檔案",
            filetypes=file_types
        )

        if filename:
            self.video_file = filename
            self.file_label.config(text=f"✅ 已選擇: {Path(filename).name}")
            self.file_label.configure(fg=JAPANESE_COLORS['success'])
            self.process_button.config(state="normal")

    def select_save_directory(self):
        """選擇截圖儲存位置"""
        directory = filedialog.askdirectory(title="選擇截圖儲存位置")

        if directory:
            self.save_directory = directory
            self.save_dir_label.config(text=f"📁 儲存到: {directory}")
            self.save_dir_label.configure(fg=JAPANESE_COLORS['text_secondary'])
        else:
            self.save_directory = None
            self.save_dir_label.config(text="📁 將儲存到程式所在目錄")
            self.save_dir_label.configure(fg=JAPANESE_COLORS['text_hint'])

    def get_frequency_range(self):
        """根據分析類型獲取頻率範圍"""
        analysis_type = self.analysis_type.get()
        if analysis_type == "high_freq":
            return {"start": 1000, "name": "高頻", "description": "機械摩擦、軸承異音等"}
        elif analysis_type == "low_freq":
            return {"start": 0, "end": 200, "name": "低頻", "description": "振動、共振、結構異音等"}
        elif analysis_type == "custom":
            start = self.custom_freq_start.get()
            end = self.custom_freq_end.get()
            # 確保起始頻率小於結束頻率
            if start >= end:
                start, end = end, start
                self.custom_freq_start.set(start)
                self.custom_freq_end.set(end)
            return {"start": start, "end": end, "name": "自訂", "description": f"{start}-{end} Hz 範圍"}
        else:  # full_freq
            return {"start": 20, "end": 20000, "name": "全頻", "description": "完整音訊頻譜"}

    def start_analysis(self):
        """開始分析（在新線程中執行）"""
        if self.is_processing:
            return

        self.is_processing = True
        self.process_button.config(state="disabled")
        self.progress.config(mode='determinate', value=0)

        # 清空之前的截圖預覽
        for widget in self.preview_scrollable_frame.winfo_children():
            widget.destroy()
        self.captured_screenshots = []

        # 在新線程中執行分析，避免GUI凍結
        thread = threading.Thread(target=self.perform_analysis)
        thread.daemon = True
        thread.start()

    def update_progress(self, value, maximum=100):
        """更新進度條"""
        self.root.after(0, lambda: self.progress.config(value=value, maximum=maximum))

    def perform_analysis(self):
        """執行完整的音訊分析"""
        try:
            total_steps = 6
            current_step = 0

            # 步驟1: 提取音訊
            current_step += 1
            self.update_progress(current_step * 100 / total_steps)
            self.update_status("🎵 正在從影片提取音訊...")
            self.extract_audio()

            # 步驟2: 頻譜分析
            current_step += 1
            self.update_progress(current_step * 100 / total_steps)
            self.update_status("📊 正在進行頻譜分析...")
            self.analyze_spectrum()

            # 步驟3: 時頻分析
            current_step += 1
            self.update_progress(current_step * 100 / total_steps)
            self.update_status("🌊 正在進行時頻分析...")
            self.analyze_time_frequency()

            # 步驟4: 截取關鍵時間點的畫面
            current_step += 1
            self.update_progress(current_step * 100 / total_steps)
            self.update_status("📸 正在截取關鍵時間點畫面...")
            self.capture_key_moments()

            # 步驟5: 繪製圖表
            current_step += 1
            self.update_progress(current_step * 100 / total_steps)
            self.update_status("🎨 正在生成精美圖表...")
            self.plot_spectrum()
            self.plot_time_frequency()

            # 步驟6: 顯示分析結果
            current_step += 1
            self.update_progress(current_step * 100 / total_steps)
            self.update_status("📋 正在整理分析結果...")
            self.display_analysis_results()
            self.display_screenshot_preview()

            # 🆕 更新視覺化總覽
            self.update_visual_summary()

            # 🆕 啟用導出按鈕
            self.root.after(0, lambda: self.export_button.config(state="normal"))

            self.update_status("✨ 分析完成！")

        except Exception as e:
            messagebox.showerror("錯誤", f"分析過程中發生錯誤：\n{str(e)}")
            self.update_status("❌ 分析失敗")
        finally:
            self.cleanup()

    def update_status(self, message):
        """更新狀態標籤（線程安全）"""
        self.root.after(0, lambda: self.status_label.config(text=message))

    def extract_audio(self):
        """從影片中提取音訊（支援時間範圍設定）"""
        clip = VideoFileClip(self.video_file)
        self.audio_duration = clip.duration

        # 🆕 如果啟用時間範圍設定，只提取指定時間段的音訊
        if self.use_time_range.get():
            start_time = max(0, self.time_start.get())
            end_time = min(self.time_end.get(), clip.duration)

            # 確保時間範圍有效
            if start_time >= end_time:
                messagebox.showwarning("警告", "開始時間必須小於結束時間，將使用完整影片。")
                audio_clip = clip.audio
            else:
                # 提取指定時間段
                audio_clip = clip.audio.subclip(start_time, end_time)
                self.audio_duration = end_time - start_time
                self.update_status(f"📌 已設定分析時間範圍: {start_time:.2f}s - {end_time:.2f}s")
        else:
            audio_clip = clip.audio

        audio_clip.write_audiofile(self.temp_audio_file, codec='pcm_s16le', verbose=False, logger=None)
        clip.close()

    def analyze_spectrum(self):
        """進行頻譜分析"""
        # 讀取WAV檔案
        self.sample_rate, data = wavfile.read(self.temp_audio_file)

        # 如果是立體聲，取左聲道
        if data.ndim > 1:
            self.audio_data = data[:, 0]
        else:
            self.audio_data = data

        # 執行FFT
        N = len(self.audio_data)
        yf = fft(self.audio_data)
        xf = fftfreq(N, 1 / self.sample_rate)

        # 只取正頻率部分
        self.xf_positive = xf[:N//2]
        yf_positive = 2.0/N * np.abs(yf[:N//2])

        # 轉換為分貝
        self.yf_db = 20 * np.log10(yf_positive, out=np.zeros_like(yf_positive),
                                   where=(yf_positive!=0))

        # 分析特定頻率範圍的噪音
        self.analyze_frequency_noise()

    def analyze_time_frequency(self):
        """進行時頻分析（支援多頻帶分析）"""
        # 設定時頻分析參數
        window_length = int(self.sample_rate * 0.1)  # 0.1秒窗口
        overlap = window_length // 2  # 50%重疊

        # 執行短時傅立葉變換 (STFT)
        frequencies, times, Zxx = signal.stft(
            self.audio_data,
            fs=self.sample_rate,
            window='hann',
            nperseg=window_length,
            noverlap=overlap
        )

        # 轉換為分貝
        magnitude_db = 20 * np.log10(np.abs(Zxx) + 1e-10)

        # 只保留人耳可聽範圍 (0-20kHz)
        freq_mask = frequencies <= 20000
        self.stft_frequencies = frequencies[freq_mask]
        self.stft_times = times
        self.stft_magnitude = magnitude_db[freq_mask, :]

        # 🆕 如果啟用多頻帶分析，分別分析各個頻帶
        if self.use_multi_band.get() and self.frequency_bands:
            self.analyze_multi_band_peaks()
        else:
            # 標準單頻帶分析
            self.analyze_time_peaks()

    def analyze_multi_band_peaks(self):
        """多頻帶峰值分析 - 分別分析各個頻帶的峰值"""
        self.multi_band_peaks = {}
        self.time_peaks = []  # 用於儲存所有頻帶的綜合峰值

        # 取得振幅閾值設定
        if self.use_amplitude_threshold.get():
            threshold = self.amplitude_threshold.get()
        else:
            threshold = -50

        # 對每個頻帶分別進行分析
        for band in self.frequency_bands:
            band_peaks = []
            band_name = band['name']

            # 設定該頻帶的頻率範圍
            freq_mask = (self.stft_frequencies >= band['start']) & (self.stft_frequencies <= band['end'])

            for i, time_point in enumerate(self.stft_times):
                # 取得該時間點的該頻帶頻譜
                target_spectrum = self.stft_magnitude[freq_mask, i]
                target_freqs = self.stft_frequencies[freq_mask]

                if len(target_spectrum) > 0:
                    max_idx = np.argmax(target_spectrum)
                    max_magnitude = target_spectrum[max_idx]
                    max_frequency = target_freqs[max_idx]

                    # 檢查是否超過閾值
                    if self.use_amplitude_threshold.get() and max_magnitude < threshold:
                        continue

                    # 找出該時間點所有超過閾值的峰值
                    peak_indices = []
                    for j in range(1, len(target_spectrum) - 1):
                        if (target_spectrum[j] > target_spectrum[j-1] and
                            target_spectrum[j] > target_spectrum[j+1] and
                            target_spectrum[j] > threshold):
                            peak_indices.append(j)

                    peak_info = {
                        'time': time_point,
                        'max_frequency': max_frequency,
                        'max_magnitude': max_magnitude,
                        'peak_count': len(peak_indices),
                        'peaks': [(target_freqs[idx], target_spectrum[idx]) for idx in peak_indices],
                        'band_name': band_name,
                        'band_color': band['color']
                    }
                    band_peaks.append(peak_info)
                    self.time_peaks.append(peak_info)  # 也加入綜合列表

            # 儲存該頻帶的峰值資料
            self.multi_band_peaks[band_name] = {
                'peaks': band_peaks,
                'color': band['color'],
                'freq_range': (band['start'], band['end'])
            }

            if band_peaks:
                # 找出該頻帶的最大峰值
                self.multi_band_peaks[band_name]['max_peak'] = max(band_peaks, key=lambda x: x['max_magnitude'])

        # 找出全域最大峰值
        if self.time_peaks:
            self.global_max_peak = max(self.time_peaks, key=lambda x: x['max_magnitude'])
        else:
            self.global_max_peak = None
            if self.use_amplitude_threshold.get():
                self.update_status(f"⚠️ 警告: 未檢測到超過 {threshold} dB 的峰值，建議降低閾值")

    def analyze_time_peaks(self):
        """分析每個時間點的峰值（支援振幅閾值設定）"""
        self.time_peaks = []
        freq_range = self.get_frequency_range()

        # 根據分析類型設定頻率範圍
        if "end" in freq_range:  # 低頻、全頻或自訂
            freq_mask = (self.stft_frequencies >= freq_range["start"]) & (self.stft_frequencies <= freq_range["end"])
        else:  # 高頻
            freq_mask = self.stft_frequencies >= freq_range["start"]

        # 🆕 取得振幅閾值設定
        if self.use_amplitude_threshold.get():
            threshold = self.amplitude_threshold.get()
        else:
            threshold = -50  # 預設閾值

        for i, time_point in enumerate(self.stft_times):
            # 取得該時間點的目標頻率範圍頻譜
            target_spectrum = self.stft_magnitude[freq_mask, i]
            target_freqs = self.stft_frequencies[freq_mask]

            # 找出該時間點的最大值
            if len(target_spectrum) > 0:
                max_idx = np.argmax(target_spectrum)
                max_magnitude = target_spectrum[max_idx]
                max_frequency = target_freqs[max_idx]

                # 🆕 如果啟用振幅閾值，檢查是否超過閾值
                if self.use_amplitude_threshold.get() and max_magnitude < threshold:
                    continue  # 跳過低於閾值的時間點

                # 找出該時間點所有超過閾值的峰值
                peak_indices = []
                for j in range(1, len(target_spectrum) - 1):
                    if (target_spectrum[j] > target_spectrum[j-1] and
                        target_spectrum[j] > target_spectrum[j+1] and
                        target_spectrum[j] > threshold):
                        peak_indices.append(j)

                peak_info = {
                    'time': time_point,
                    'max_frequency': max_frequency,
                    'max_magnitude': max_magnitude,
                    'peak_count': len(peak_indices),
                    'peaks': [(target_freqs[idx], target_spectrum[idx]) for idx in peak_indices]
                }
                self.time_peaks.append(peak_info)

        # 找出全域最大峰值
        if self.time_peaks:
            self.global_max_peak = max(self.time_peaks, key=lambda x: x['max_magnitude'])
        else:
            self.global_max_peak = None
            # 如果沒有峰值，可能是閾值設定過高
            if self.use_amplitude_threshold.get():
                self.update_status(f"⚠️ 警告: 未檢測到超過 {threshold} dB 的峰值，建議降低閾值")

    def capture_key_moments(self):
        """截取關鍵時間點的畫面和音訊（支援時間範圍調整和音訊節錄）"""
        if not self.time_peaks:
            return

        # 獲取要截圖的數量
        screenshot_count = min(self.screenshot_count.get(), len(self.time_peaks))

        # 按振幅排序，選取前N個峰值
        sorted_peaks = sorted(self.time_peaks, key=lambda x: x['max_magnitude'], reverse=True)[:screenshot_count]

        # 設定儲存目錄
        if self.save_directory:
            save_dir = Path(self.save_directory)
        else:
            save_dir = Path.cwd()

        # 創建以當前時間命名的子目錄
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        video_name = Path(self.video_file).stem

        # 🆕 根據是否使用多頻帶分析來命名
        if self.use_multi_band.get() and self.frequency_bands:
            analysis_type_name = "多頻帶"
        else:
            analysis_type_name = self.get_frequency_range()["name"]

        screenshot_dir = save_dir / f"{video_name}_{analysis_type_name}分析_screenshots_{timestamp}"
        screenshot_dir.mkdir(exist_ok=True)

        # 🆕 創建音訊子目錄
        if self.enable_audio_extract.get():
            audio_dir = screenshot_dir / "audio_clips"
            audio_dir.mkdir(exist_ok=True)

        # 開啟影片進行截圖
        clip = VideoFileClip(self.video_file)

        self.captured_screenshots = []
        self.extracted_audio_clips = []

        # 如果啟用時間範圍，需要調整截圖時間點
        time_offset = 0
        if self.use_time_range.get():
            time_offset = self.time_start.get()

        for i, peak in enumerate(sorted_peaks):
            try:
                # 計算影片中的實際時間點（考慮時間偏移）
                analysis_time = peak['time']
                actual_time = analysis_time + time_offset

                # 確保時間點在影片範圍內
                actual_time = max(0, min(actual_time, clip.duration - 0.1))

                # 獲取畫面
                frame = clip.get_frame(actual_time)

                # 轉換為PIL圖像
                image = Image.fromarray(frame)

                # 🆕 確定頻帶資訊（用於檔名）
                band_info = ""
                if 'band_name' in peak:
                    band_info = f"_{peak['band_name']}"

                # 生成檔案名稱
                filename = f"{analysis_type_name}_peak_{i+1:02d}{band_info}_time_{actual_time:.2f}s_freq_{peak['max_frequency']:.0f}Hz_mag_{peak['max_magnitude']:.1f}dB.jpg"
                filepath = screenshot_dir / filename

                # 儲存圖像
                image.save(filepath, "JPEG", quality=90)

                # 🆕 節錄音訊片段
                audio_filepath = None
                if self.enable_audio_extract.get():
                    try:
                        audio_filepath = self.extract_audio_clip(
                            clip, actual_time, audio_dir, i+1, peak, band_info
                        )
                    except Exception as e:
                        print(f"音訊節錄失敗 {actual_time:.2f}s: {e}")

                # 記錄截圖資訊
                screenshot_info = {
                    'rank': i + 1,
                    'time': actual_time,
                    'analysis_time': analysis_time,
                    'frequency': peak['max_frequency'],
                    'magnitude': peak['max_magnitude'],
                    'filepath': filepath,
                    'image': image.copy(),
                    'audio_filepath': audio_filepath,  # 🆕 音訊檔案路徑
                    'band_name': peak.get('band_name', ''),  # 🆕 頻帶名稱
                    'band_color': peak.get('band_color', JAPANESE_COLORS['accent_primary'])  # 🆕 頻帶顏色
                }
                self.captured_screenshots.append(screenshot_info)

                # 更新進度
                progress = 50 + (i + 1) / screenshot_count * 30
                self.update_progress(progress)

            except Exception as e:
                print(f"截圖失敗 {actual_time:.2f}s: {e}")
                continue

        clip.close()

        # 儲存截圖資訊檔案
        self.save_analysis_info(screenshot_dir, analysis_type_name)

    def extract_audio_clip(self, clip, peak_time, audio_dir, rank, peak_info, band_info):
        """節錄音訊片段"""
        # 計算音訊片段的起始和結束時間
        duration_before = self.audio_duration_before.get()
        duration_after = self.audio_duration_after.get()

        start_time = max(0, peak_time - duration_before)
        end_time = min(clip.duration, peak_time + duration_after)

        # 提取音訊片段
        audio_clip = clip.audio.subclip(start_time, end_time)

        # 生成音訊檔案名稱
        audio_filename = f"audio_clip_{rank:02d}{band_info}_time_{peak_time:.2f}s_freq_{peak_info['max_frequency']:.0f}Hz.wav"
        audio_filepath = audio_dir / audio_filename

        # 儲存為WAV格式
        audio_clip.write_audiofile(str(audio_filepath), codec='pcm_s16le',
                                   fps=44100, verbose=False, logger=None)

        # 記錄音訊片段資訊
        self.extracted_audio_clips.append({
            'filepath': audio_filepath,
            'start_time': start_time,
            'end_time': end_time,
            'peak_time': peak_time,
            'duration': end_time - start_time
        })

        return audio_filepath

    def save_analysis_info(self, screenshot_dir, analysis_type_name):
        """儲存分析資訊檔案"""
        info_file = screenshot_dir / "screenshot_info.txt"
        with open(info_file, 'w', encoding='utf-8') as f:
            f.write(f"影片檔案: {Path(self.video_file).name}\n")
            f.write(f"分析類型: {analysis_type_name}分析\n")

            # 記錄分析設定
            if self.use_multi_band.get() and self.frequency_bands:
                f.write(f"多頻帶設定:\n")
                for band in self.frequency_bands:
                    f.write(f"  - {band['name']}: {band['start']}-{band['end']} Hz\n")
            else:
                freq_range = self.get_frequency_range()
                if "end" in freq_range:
                    f.write(f"頻率範圍: {freq_range['start']}-{freq_range['end']} Hz\n")
                else:
                    f.write(f"頻率範圍: >{freq_range['start']} Hz\n")

            if self.use_time_range.get():
                f.write(f"分析時間段: {self.time_start.get():.2f}s - {self.time_end.get():.2f}s\n")

            if self.use_amplitude_threshold.get():
                f.write(f"振幅閾值: {self.amplitude_threshold.get()} dB\n")

            if self.enable_audio_extract.get():
                f.write(f"音訊節錄: 峰值前 {self.audio_duration_before.get():.1f}s + 峰值後 {self.audio_duration_after.get():.1f}s\n")

            f.write(f"分析時間: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"總共截圖: {len(self.captured_screenshots)} 張\n")
            if self.enable_audio_extract.get():
                f.write(f"總共節錄音訊: {len(self.extracted_audio_clips)} 個\n")
            f.write("\n")
            f.write("截圖詳細資訊:\n")
            f.write("="*60 + "\n")

            for screenshot in self.captured_screenshots:
                f.write(f"排名: {screenshot['rank']}\n")
                if screenshot['band_name']:
                    f.write(f"頻帶: {screenshot['band_name']}\n")
                f.write(f"影片時間: {screenshot['time']:.2f} 秒\n")
                if self.use_time_range.get():
                    f.write(f"分析時間: {screenshot['analysis_time']:.2f} 秒\n")
                f.write(f"頻率: {screenshot['frequency']:.1f} Hz\n")
                f.write(f"振幅: {screenshot['magnitude']:.1f} dB\n")
                f.write(f"圖片檔案: {screenshot['filepath'].name}\n")
                if screenshot['audio_filepath']:
                    f.write(f"音訊檔案: {screenshot['audio_filepath'].name}\n")
                f.write("-" * 40 + "\n")

    def display_screenshot_preview(self):
        """顯示截圖預覽 - 增強邊框效果並支援音訊播放"""
        def update_preview():
            # 清空現有預覽
            for widget in self.preview_scrollable_frame.winfo_children():
                widget.destroy()

            if not self.captured_screenshots:
                no_screenshot_frame = tk.Frame(self.preview_scrollable_frame,
                                             bg=JAPANESE_COLORS['bg_card'],
                                             relief='solid', borderwidth=2,
                                             highlightbackground=JAPANESE_COLORS['border'])
                no_screenshot_frame.pack(side=tk.LEFT, padx=20, pady=20)

                no_screenshot_label = tk.Label(no_screenshot_frame,
                                              text="📷 沒有截圖\n請確保影片檔案有效",
                                              font=self.fonts['body'],
                                              fg=JAPANESE_COLORS['text_hint'],
                                              bg=JAPANESE_COLORS['bg_card'],
                                              justify=tk.CENTER)
                no_screenshot_label.pack(padx=30, pady=30)
                return

            # 顯示截圖預覽
            for i, screenshot in enumerate(self.captured_screenshots):
                # 創建預覽框架 - 增強邊框效果，並使用頻帶顏色
                border_color = screenshot.get('band_color', JAPANESE_COLORS['border_strong'])
                preview_frame = tk.Frame(self.preview_scrollable_frame,
                                       bg=JAPANESE_COLORS['bg_card'],
                                       relief="solid", borderwidth=3,
                                       highlightbackground=border_color,
                                       highlightthickness=2)
                preview_frame.pack(side=tk.LEFT, padx=8, pady=8)

                # 🆕 頻帶標籤（如果有）
                if screenshot.get('band_name'):
                    band_label_frame = tk.Frame(preview_frame, bg=border_color)
                    band_label_frame.pack(fill=tk.X, padx=8, pady=(8, 0))

                    band_label = tk.Label(band_label_frame,
                                        text=f"🎼 {screenshot['band_name']}",
                                        font=self.fonts['small'],
                                        fg='white',
                                        bg=border_color)
                    band_label.pack(pady=3)

                # 縮小圖像用於預覽
                image = screenshot['image']
                height = 90
                width = int(image.width * height / image.height)
                thumbnail = image.resize((width, height), Image.Resampling.LANCZOS)

                # 轉換為tkinter可用的圖像
                from PIL import ImageTk
                photo = ImageTk.PhotoImage(thumbnail)

                # 圖像容器 - 添加內部邊框
                img_container = tk.Frame(preview_frame, bg=JAPANESE_COLORS['border'],
                                       relief='solid', borderwidth=1)
                img_container.pack(padx=8, pady=(8, 5))

                # 顯示縮圖
                img_label = tk.Label(img_container, image=photo,
                                   bg=JAPANESE_COLORS['bg_card'], borderwidth=0)
                img_label.image = photo  # 保持引用
                img_label.pack(padx=2, pady=2)

                # 資訊容器 - 添加背景邊框
                info_container = tk.Frame(preview_frame, bg=JAPANESE_COLORS['bg_secondary'],
                                        relief='solid', borderwidth=1)
                info_container.pack(padx=8, pady=(0, 5), fill=tk.X)

                # 顯示資訊
                info_text = f"#{screenshot['rank']} - {screenshot['time']:.1f}s\n{screenshot['frequency']:.0f}Hz ({screenshot['magnitude']:.1f}dB)"
                info_label = tk.Label(info_container, text=info_text,
                                    font=self.fonts['small'], justify=tk.CENTER,
                                    bg=JAPANESE_COLORS['bg_secondary'],
                                    fg=JAPANESE_COLORS['text_primary'])
                info_label.pack(padx=8, pady=8)

                # 🆕 音訊播放按鈕（如果有音訊檔案）
                if screenshot.get('audio_filepath'):
                    audio_btn_frame = tk.Frame(preview_frame, bg=JAPANESE_COLORS['bg_card'])
                    audio_btn_frame.pack(padx=8, pady=(0, 8), fill=tk.X)

                    play_btn = ttk.Button(audio_btn_frame, text="▶️ 播放音訊",
                                        command=lambda fp=screenshot['audio_filepath']: self.play_audio(fp),
                                        style='Secondary.TButton')
                    play_btn.pack(fill=tk.X)

                # 添加點擊事件打開完整圖像
                def open_full_image(filepath=screenshot['filepath']):
                    try:
                        if platform.system() == "Windows":
                            os.startfile(filepath)
                        elif platform.system() == "Darwin":  # macOS
                            os.system(f"open '{filepath}'")
                        else:  # Linux
                            os.system(f"xdg-open '{filepath}'")
                    except Exception as e:
                        messagebox.showerror("錯誤", f"無法開啟圖像: {e}")

                img_label.bind("<Button-1>", lambda e, fp=screenshot['filepath']: open_full_image(fp))
                img_label.config(cursor="hand2")

                # 添加懸停效果
                def on_enter(e, frame=preview_frame, color=border_color):
                    frame.configure(highlightbackground=JAPANESE_COLORS['accent_primary'])
                def on_leave(e, frame=preview_frame, color=border_color):
                    frame.configure(highlightbackground=color)

                preview_frame.bind("<Enter>", on_enter)
                preview_frame.bind("<Leave>", on_leave)
                img_label.bind("<Enter>", on_enter)
                img_label.bind("<Leave>", on_leave)
                info_label.bind("<Enter>", on_enter)
                info_label.bind("<Leave>", on_leave)

        self.root.after(0, update_preview)

    def play_audio(self, audio_filepath):
        """播放音訊檔案"""
        try:
            # 使用系統預設播放器播放音訊
            if platform.system() == "Windows":
                os.startfile(audio_filepath)
            elif platform.system() == "Darwin":  # macOS
                os.system(f"afplay '{audio_filepath}' &")
            else:  # Linux
                os.system(f"paplay '{audio_filepath}' &")
        except Exception as e:
            # 如果系統播放器失敗，嘗試使用其他方法
            try:
                # 嘗試使用默認應用程式打開
                import subprocess
                if platform.system() == "Windows":
                    subprocess.Popen(['start', '', str(audio_filepath)], shell=True)
                elif platform.system() == "Darwin":
                    subprocess.Popen(['open', str(audio_filepath)])
                else:
                    subprocess.Popen(['xdg-open', str(audio_filepath)])
            except:
                messagebox.showinfo("提示",
                                  f"無法自動播放音訊。\n請手動開啟檔案：\n{audio_filepath.name}\n\n位置：{audio_filepath.parent}")

    def analyze_frequency_noise(self):
        """分析特定頻率範圍的噪音"""
        freq_range = self.get_frequency_range()

        # 根據分析類型設定頻率範圍
        if "end" in freq_range:  # 低頻或全頻
            freq_mask = (self.xf_positive >= freq_range["start"]) & (self.xf_positive <= freq_range["end"])
        else:  # 高頻
            freq_mask = self.xf_positive >= freq_range["start"]

        target_freqs = self.xf_positive[freq_mask]
        target_magnitudes = self.yf_db[freq_mask]

        # 找出峰值（簡單的峰值檢測）
        peaks = []
        threshold = -60  # dB閾值

        for i in range(1, len(target_magnitudes) - 1):
            if (target_magnitudes[i] > target_magnitudes[i-1] and
                target_magnitudes[i] > target_magnitudes[i+1] and
                target_magnitudes[i] > threshold):
                peaks.append({
                    'frequency': target_freqs[i],
                    'magnitude': target_magnitudes[i]
                })

        # 按振幅排序，取前10個最強的峰值
        self.freq_peaks = sorted(peaks, key=lambda x: x['magnitude'], reverse=True)[:10]

        # 統計資訊
        self.analysis_stats = {
            'total_peaks': len(peaks),
            'max_magnitude': max(target_magnitudes) if len(target_magnitudes) > 0 else 0,
            'avg_magnitude': np.mean(target_magnitudes) if len(target_magnitudes) > 0 else 0,
            'freq_range': freq_range
        }

    def plot_spectrum(self):
        """繪製頻譜圖"""
        def update_plot():
            self.ax1.clear()
            self.ax1.plot(self.xf_positive, self.yf_db, color=JAPANESE_COLORS['accent_primary'],
                         linewidth=1.5, alpha=0.8)

            freq_range = self.get_frequency_range()

            # 根據分析類型添加標記線
            if freq_range["name"] == "高頻":
                self.ax1.axvline(x=freq_range["start"], color=JAPANESE_COLORS['warning'],
                               linestyle='--', linewidth=2,
                               label=f'高頻區域 (>{freq_range["start"]} Hz)', alpha=0.8)
            elif freq_range["name"] == "低頻":
                self.ax1.axvline(x=freq_range["end"], color=JAPANESE_COLORS['warning'],
                               linestyle='--', linewidth=2,
                               label=f'低頻區域 (<{freq_range["end"]} Hz)', alpha=0.8)
            else:  # 全頻
                self.ax1.axvspan(freq_range["start"], freq_range["end"],
                               color=JAPANESE_COLORS['accent_secondary'], alpha=0.1,
                               label=f'分析範圍 ({freq_range["start"]}-{freq_range["end"]} Hz)')

            self.ax1.set_title(f"音訊{freq_range['name']}頻譜分析 (dB)",
                             fontsize=14, color=JAPANESE_COLORS['text_primary'], pad=20)
            self.ax1.set_xlabel("頻率 (Hz)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
            self.ax1.set_ylabel("振幅 (dB)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
            self.ax1.grid(True, which="both", linestyle='--', linewidth=0.5, color=JAPANESE_COLORS['divider'])
            self.ax1.legend(prop={'size': 10})
            self.ax1.set_xlim(0, 20000)
            self.ax1.set_facecolor(JAPANESE_COLORS['bg_secondary'])

            # 標記峰值
            for peak in self.freq_peaks[:5]:  # 只標記前5個最強峰值
                self.ax1.plot(peak['frequency'], peak['magnitude'], 'o',
                             color=JAPANESE_COLORS['error'], markersize=8, alpha=0.8,
                             markeredgecolor='white', markeredgewidth=1)
                self.ax1.annotate(f"{peak['frequency']:.0f}Hz",
                               xy=(peak['frequency'], peak['magnitude']),
                               xytext=(5, 5), textcoords='offset points',
                               fontsize=8, ha='left', color=JAPANESE_COLORS['text_primary'])

            self.canvas1.draw()

        self.root.after(0, update_plot)

    def plot_time_frequency(self):
        """繪製時頻分析圖（支援多頻帶對比顯示）"""
        def update_plot():
            # 清空圖表
            self.ax2.clear()
            self.ax3.clear()

            freq_range = self.get_frequency_range()

            # 繪製時頻圖（頻譜圖）
            im = self.ax2.imshow(
                self.stft_magnitude,
                aspect='auto',
                origin='lower',
                extent=[self.stft_times[0], self.stft_times[-1],
                       self.stft_frequencies[0], self.stft_frequencies[-1]],
                cmap='viridis',
                vmin=-80, vmax=0
            )

            # 🆕 多頻帶分析顯示
            if self.use_multi_band.get() and self.frequency_bands:
                self.ax2.set_title(f"時頻分析圖 - 多頻帶對比分析",
                                 fontsize=14, color=JAPANESE_COLORS['text_primary'], pad=15)

                # 標記各個頻帶範圍
                for band in self.frequency_bands:
                    # 在頻率軸上標記頻帶範圍
                    self.ax2.axhspan(band['start'], band['end'],
                                   color=band['color'], alpha=0.1, linewidth=0)
                    # 標記頻帶邊界
                    self.ax2.axhline(y=band['start'], color=band['color'],
                                   linestyle='--', linewidth=1.5, alpha=0.6)
                    self.ax2.axhline(y=band['end'], color=band['color'],
                                   linestyle='--', linewidth=1.5, alpha=0.6)
                    # 標註頻帶名稱
                    mid_freq = (band['start'] + band['end']) / 2
                    self.ax2.text(self.stft_times[-1] * 1.01, mid_freq, band['name'],
                               color=band['color'], fontsize=8, va='center', weight='bold')
            else:
                self.ax2.set_title(f"時頻分析圖 - {freq_range['name']}頻譜分析",
                                 fontsize=14, color=JAPANESE_COLORS['text_primary'], pad=15)

                # 根據分析類型添加頻率範圍標記
                if freq_range["name"] == "高頻":
                    self.ax2.axhline(y=freq_range["start"], color=JAPANESE_COLORS['warning'],
                                   linestyle='--', linewidth=2, alpha=0.8)
                elif "end" in freq_range:
                    self.ax2.axhline(y=freq_range["start"], color=JAPANESE_COLORS['warning'],
                                   linestyle='--', linewidth=2, alpha=0.8)
                    self.ax2.axhline(y=freq_range["end"], color=JAPANESE_COLORS['warning'],
                                   linestyle='--', linewidth=2, alpha=0.8)

            self.ax2.set_xlabel("時間 (秒)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
            self.ax2.set_ylabel("頻率 (Hz)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
            self.ax2.set_facecolor(JAPANESE_COLORS['bg_secondary'])

            # 添加顏色條
            cbar = self.fig2.colorbar(im, ax=self.ax2)
            cbar.set_label('振幅 (dB)', fontsize=10, color=JAPANESE_COLORS['text_primary'])

            # 標記截圖時間點（按頻帶顏色分組）
            for i, screenshot in enumerate(self.captured_screenshots[:10]):
                color = screenshot.get('band_color', JAPANESE_COLORS['accent_secondary'])
                self.ax2.axvline(x=screenshot['time'], color=color,
                               linestyle='-', linewidth=2.5, alpha=0.8)
                if i < 5:
                    label = f"#{screenshot['rank']}"
                    if screenshot.get('band_name'):
                        label += f"\n{screenshot['band_name']}"
                    self.ax2.annotate(label,
                                    xy=(screenshot['time'], screenshot['frequency']),
                                    xytext=(5, 5), textcoords='offset points',
                                    fontsize=7, color=color, weight='bold')

            # 標記全域最大峰值位置
            if self.global_max_peak:
                self.ax2.plot(self.global_max_peak['time'],
                             self.global_max_peak['max_frequency'],
                             '*', color=JAPANESE_COLORS['error'], markersize=20,
                             markeredgecolor='white', markeredgewidth=2)

            # 🆕 繪製多頻帶對比的時間-峰值振幅圖
            if self.use_multi_band.get() and self.multi_band_peaks:
                # 為每個頻帶繪製振幅曲線
                for band_name, band_data in self.multi_band_peaks.items():
                    peaks = band_data['peaks']
                    if peaks:
                        times = [p['time'] for p in peaks]
                        magnitudes = [p['max_magnitude'] for p in peaks]
                        color = band_data['color']

                        # 繪製該頻帶的振幅線
                        self.ax3.plot(times, magnitudes, color=color,
                                    linewidth=2, label=f'{band_name} 最大振幅', alpha=0.8)
                        self.ax3.fill_between(times, magnitudes, color=color, alpha=0.15)

                        # 標記該頻帶的最高峰值
                        max_peak = band_data['max_peak']
                        self.ax3.plot(max_peak['time'], max_peak['max_magnitude'],
                                    'o', color=color, markersize=10,
                                    markeredgecolor='white', markeredgewidth=2)

                self.ax3.set_title("多頻帶振幅對比 - 各頻帶最大振幅變化",
                                 fontsize=14, color=JAPANESE_COLORS['text_primary'], pad=15)
            else:
                # 標準單頻帶顯示
                times = [peak['time'] for peak in self.time_peaks]
                max_magnitudes = [peak['max_magnitude'] for peak in self.time_peaks]

                self.ax3.plot(times, max_magnitudes, color=JAPANESE_COLORS['accent_primary'],
                            linewidth=2, label=f'{freq_range["name"]}頻最大振幅', alpha=0.8)
                self.ax3.fill_between(times, max_magnitudes, color=JAPANESE_COLORS['accent_primary'], alpha=0.2)

                self.ax3.set_title(f"各時間點{freq_range['name']}頻最大振幅變化",
                                 fontsize=14, color=JAPANESE_COLORS['text_primary'], pad=15)

            # 標記截圖時間點
            for screenshot in self.captured_screenshots:
                color = screenshot.get('band_color', JAPANESE_COLORS['error'])
                self.ax3.axvline(x=screenshot['time'], color=color,
                               linestyle='-', linewidth=2, alpha=0.6)
                self.ax3.plot(screenshot['time'], screenshot['magnitude'],
                             'o', color=color, markersize=10,
                             markeredgecolor='white', markeredgewidth=2)

            # 標記全域最大峰值
            if self.global_max_peak:
                self.ax3.plot(self.global_max_peak['time'],
                             self.global_max_peak['max_magnitude'],
                             '*', color=JAPANESE_COLORS['error'], markersize=20,
                             markeredgecolor='white', markeredgewidth=2)

            self.ax3.set_xlabel("時間 (秒)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
            self.ax3.set_ylabel("最大振幅 (dB)", fontsize=12, color=JAPANESE_COLORS['text_primary'])
            self.ax3.grid(True, linestyle='--', linewidth=0.5, color=JAPANESE_COLORS['divider'])
            self.ax3.legend(loc='best', prop={'size': 9})
            self.ax3.set_facecolor(JAPANESE_COLORS['bg_secondary'])

            self.fig2.tight_layout(pad=3.0)
            self.canvas2.draw()

        self.root.after(0, update_plot)

    def display_analysis_results(self):
        """顯示分析結果"""
        def update_results():
            self.result_text.delete(1.0, tk.END)

            freq_range = self.get_frequency_range()

            # 基本資訊
            result_text = f"🎵 === 音訊{freq_range['name']}頻譜分析結果 (含時間維度及截圖) === 🎵\n\n"
            result_text += f"📁 影片檔案: {Path(self.video_file).name}\n"
            result_text += f"🔧 分析類型: {freq_range['name']}頻分析 ({freq_range['description']})\n"

            # 🆕 顯示分析設定
            if freq_range["name"] == "高頻":
                result_text += f"🔊 分析頻率範圍: >{freq_range['start']} Hz (高頻區域)\n"
            elif "end" in freq_range:
                result_text += f"🔊 分析頻率範圍: {freq_range['start']}-{freq_range['end']} Hz\n"

            if self.use_time_range.get():
                result_text += f"⏰ 分析時間範圍: {self.time_start.get():.2f}s - {self.time_end.get():.2f}s\n"
                result_text += f"   (總長度: {self.audio_duration:.2f} 秒)\n"
            else:
                result_text += f"⏱️ 影片總長度: {self.audio_duration:.2f} 秒\n"

            if self.use_amplitude_threshold.get():
                result_text += f"📊 振幅閾值: {self.amplitude_threshold.get()} dB (只分析超過此強度的聲音)\n"

            result_text += f"🎚️ 音訊取樣率: {self.sample_rate} Hz\n\n"

            # 截圖資訊
            result_text += "📸 === 截圖資訊 === 📸\n"
            if self.captured_screenshots:
                result_text += f"✨ 共截取 {len(self.captured_screenshots)} 張關鍵時間點畫面\n"
                screenshot_dir = self.captured_screenshots[0]['filepath'].parent
                result_text += f"📁 儲存位置: {screenshot_dir}\n"
                result_text += "🖱️ 點擊下方預覽圖像可開啟完整圖片\n\n"

                result_text += "📋 截圖詳細列表:\n"
                for screenshot in self.captured_screenshots:
                    result_text += f"#{screenshot['rank']:2d}. ⏰{screenshot['time']:6.2f}秒 - "
                    result_text += f"🎵{screenshot['frequency']:7.1f}Hz - 📊{screenshot['magnitude']:6.2f}dB\n"
                    result_text += f"     📄 檔案: {screenshot['filepath'].name}\n"
                result_text += "\n"
            else:
                result_text += "❌ 未能成功截取畫面\n\n"

            # 整體統計資訊
            result_text += f"📈 === {freq_range['name']}頻統計資訊 === 📈\n"
            result_text += f"🔍 檢測到的{freq_range['name']}頻峰值數量: {self.analysis_stats['total_peaks']}\n"
            result_text += f"📊 {freq_range['name']}頻區域最大振幅: {self.analysis_stats['max_magnitude']:.2f} dB\n"
            result_text += f"📊 {freq_range['name']}頻區域平均振幅: {self.analysis_stats['avg_magnitude']:.2f} dB\n\n"

            # 時間分析結果
            result_text += "⏰ === 時間維度分析結果 === ⏰\n"
            if self.global_max_peak:
                result_text += f"🎯 全域最大峰值出現時間: {self.global_max_peak['time']:.2f} 秒\n"
                result_text += f"   🎵 對應頻率: {self.global_max_peak['max_frequency']:.1f} Hz\n"
                result_text += f"   📊 振幅大小: {self.global_max_peak['max_magnitude']:.2f} dB\n"
                result_text += f"   🔢 該時間點峰值數量: {self.global_max_peak['peak_count']}\n\n"

            # 主要峰值頻率
            if self.freq_peaks:
                result_text += f"🎵 === 主要{freq_range['name']}頻峰值 (整體平均) === 🎵\n"
                for i, peak in enumerate(self.freq_peaks, 1):
                    result_text += f"{i:2d}. 🎵 {peak['frequency']:8.1f} Hz - 📊 {peak['magnitude']:6.2f} dB\n"
            else:
                result_text += f"🎵 === {freq_range['name']}頻峰值 === 🎵\n"
                result_text += f"✅ 未檢測到明顯的{freq_range['name']}頻峰值\n"

            result_text += "\n"

            # 分析建議
            result_text += "💡 === 分析重點提示 === 💡\n"
            result_text += "1. 🖼️ 查看截圖預覽區域，了解異常音訊產生時的機械動作\n"
            result_text += "2. 📊 時頻分析圖中的綠色垂直線標示了截圖時間點\n"
            result_text += "3. ⭐ 紅色星號標記了全域最大峰值的時間和頻率位置\n"
            result_text += "4. 📈 振幅變化圖中的紅點顯示了所有截圖時間點\n"
            result_text += "5. 🔍 通過對比截圖與頻譜，可以找出產生異常音訊的具體機械動作\n"
            result_text += "6. 📋 詳細的截圖資訊已儲存為文字檔案，便於後續分析\n\n"

            # 根據分析類型提供特定建議
            if freq_range["name"] == "高頻":
                result_text += "🔧 === 高頻分析特殊建議 === 🔧\n"
                result_text += "• 關注軸承、齒輪等旋轉部件的異音\n"
                result_text += "• 檢查電機或馬達的高頻電磁噪音\n"
                result_text += "• 注意金屬摩擦或碰撞產生的尖銳聲音\n\n"
            elif freq_range["name"] == "低頻":
                result_text += "🔧 === 低頻分析特殊建議 === 🔧\n"
                result_text += "• 關注結構振動或共振問題\n"
                result_text += "• 檢查大型部件的不平衡或鬆動\n"
                result_text += "• 注意低速旋轉部件的週期性異音\n\n"

            # 健康狀態評估
            result_text += "🏥 === 設備狀態評估 === 🏥\n"
            if self.global_max_peak:
                if self.global_max_peak['max_magnitude'] > -30:
                    result_text += "🔴 警告: 檢測到非常高的峰值，建議立即檢查設備\n"
                    result_text += "   🔧 結合截圖分析該時間點的機械動作\n"
                elif self.global_max_peak['max_magnitude'] > -40:
                    result_text += "🟡 注意: 檢測到較高的峰值，建議密切監控\n"
                    result_text += "   🔧 建議檢查截圖中對應的機械部件\n"
                else:
                    result_text += "🟢 良好: 峰值在正常範圍內\n"

                # 分析峰值分佈的時間特性
                if len(self.time_peaks) > 0:
                    high_magnitude_count = sum(1 for p in self.time_peaks if p['max_magnitude'] > -40)
                    high_ratio = high_magnitude_count / len(self.time_peaks)

                    if high_ratio > 0.3:
                        result_text += "⚠️ 高振幅時間點比例較高，設備可能存在持續性問題\n"
                        result_text += "   🔧 建議檢查截圖中重複出現的機械動作\n"
                    elif high_ratio > 0.1:
                        result_text += "ℹ️ 偶有高振幅時間點，建議定期監控\n"
                    else:
                        result_text += "✅ 高振幅時間點比例較低，設備狀況良好\n"
            else:
                result_text += "✅ 未檢測到明顯異常，設備狀況良好\n"

            self.result_text.insert(tk.END, result_text)

        self.root.after(0, update_results)

    def export_to_excel(self):
        """導出分析結果到 Excel 檔案"""
        if not hasattr(self, 'time_peaks') or not self.time_peaks:
            messagebox.showwarning("警告", "尚無分析資料可導出！\n請先完成分析。")
            return

        # 選擇儲存位置
        default_filename = f"音訊分析報告_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = filedialog.asksaveasfilename(
            title="儲存 Excel 分析報告",
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*")]
        )

        if not filepath:
            return

        try:
            self.update_status("📊 正在生成 Excel 報告...")

            # 創建 Excel 工作簿
            wb = openpyxl.Workbook()

            # 移除預設工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 創建各個工作表
            self.create_summary_sheet(wb)
            self.create_peaks_sheet(wb)
            self.create_spectrum_sheet(wb)
            self.create_timefreq_sheet(wb)
            self.create_screenshots_sheet(wb)

            if self.use_multi_band.get() and self.multi_band_peaks:
                self.create_multiband_sheet(wb)

            # 儲存檔案
            wb.save(filepath)

            self.update_status("✅ Excel 報告已成功導出！")

            # 詢問是否開啟檔案
            if messagebox.askyesno("完成", f"Excel 報告已成功導出！\n\n檔案位置：\n{filepath}\n\n是否要開啟檔案？"):
                try:
                    if platform.system() == "Windows":
                        os.startfile(filepath)
                    elif platform.system() == "Darwin":
                        os.system(f"open '{filepath}'")
                    else:
                        os.system(f"xdg-open '{filepath}'")
                except:
                    messagebox.showinfo("提示", f"請手動開啟檔案：\n{filepath}")

        except Exception as e:
            messagebox.showerror("錯誤", f"導出 Excel 時發生錯誤：\n{str(e)}")
            self.update_status("❌ Excel 導出失敗")

    def create_summary_sheet(self, wb):
        """創建分析摘要工作表"""
        ws = wb.create_sheet("📊 分析摘要", 0)

        # 設定欄寬
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

        # 標題樣式
        title_font = Font(size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")

        # 標題
        ws.merge_cells('A1:B1')
        ws['A1'] = "音訊頻譜分析報告"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = title_alignment
        ws.row_dimensions[1].height = 30

        # 基本資訊
        row = 3
        header_font = Font(bold=True, size=11)

        info_data = [
            ("影片檔案", Path(self.video_file).name),
            ("分析時間", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("影片長度", f"{self.audio_duration:.2f} 秒"),
            ("音訊取樣率", f"{self.sample_rate} Hz"),
            ("", ""),
            ("分析設定", ""),
        ]

        # 分析類型
        if self.use_multi_band.get() and self.frequency_bands:
            info_data.append(("分析模式", "多頻帶同步分析"))
            for band in self.frequency_bands:
                info_data.append(("  - " + band['name'], f"{band['start']}-{band['end']} Hz"))
        else:
            freq_range = self.get_frequency_range()
            info_data.append(("分析模式", f"{freq_range['name']}頻分析"))
            if "end" in freq_range:
                info_data.append(("頻率範圍", f"{freq_range['start']}-{freq_range['end']} Hz"))
            else:
                info_data.append(("頻率範圍", f">{freq_range['start']} Hz"))

        if self.use_time_range.get():
            info_data.append(("分析時間段", f"{self.time_start.get():.2f}s - {self.time_end.get():.2f}s"))

        if self.use_amplitude_threshold.get():
            info_data.append(("振幅閾值", f"{self.amplitude_threshold.get()} dB"))

        info_data.extend([
            ("", ""),
            ("分析結果", ""),
            ("檢測峰值數量", f"{len(self.time_peaks)} 個"),
            ("截圖數量", f"{len(self.captured_screenshots)} 張"),
        ])

        if self.enable_audio_extract.get():
            info_data.append(("音訊片段數量", f"{len(self.extracted_audio_clips)} 個"))

        # 全域最大峰值
        if self.global_max_peak:
            info_data.extend([
                ("", ""),
                ("最大峰值資訊", ""),
                ("出現時間", f"{self.global_max_peak['time']:.2f} 秒"),
                ("峰值頻率", f"{self.global_max_peak['max_frequency']:.1f} Hz"),
                ("峰值振幅", f"{self.global_max_peak['max_magnitude']:.2f} dB"),
            ])

        # 寫入資料
        for item, value in info_data:
            if item == "" and value == "":
                row += 1
                continue

            ws[f'A{row}'] = item
            ws[f'B{row}'] = value

            if value == "" or item in ["分析設定", "分析結果", "最大峰值資訊"]:
                ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
                ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                ws.merge_cells(f'A{row}:B{row}')
            else:
                ws[f'A{row}'].font = header_font

            row += 1

        # 添加邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border

    def create_peaks_sheet(self, wb):
        """創建峰值資料工作表"""
        ws = wb.create_sheet("🔍 峰值資料")

        # 標題
        headers = ["排名", "時間 (秒)", "頻率 (Hz)", "振幅 (dB)", "峰值數量"]

        if self.use_multi_band.get() and self.frequency_bands:
            headers.insert(1, "頻帶")

        # 寫入標題
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 寫入資料
        sorted_peaks = sorted(self.time_peaks, key=lambda x: x['max_magnitude'], reverse=True)

        for idx, peak in enumerate(sorted_peaks, 1):
            row = idx + 1
            col = 1

            ws.cell(row=row, column=col, value=idx)
            col += 1

            if self.use_multi_band.get() and 'band_name' in peak:
                ws.cell(row=row, column=col, value=peak['band_name'])
                col += 1

            ws.cell(row=row, column=col, value=round(peak['time'], 2))
            col += 1
            ws.cell(row=row, column=col, value=round(peak['max_frequency'], 1))
            col += 1
            ws.cell(row=row, column=col, value=round(peak['max_magnitude'], 2))
            col += 1
            ws.cell(row=row, column=col, value=peak['peak_count'])

        # 自動調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 添加圖表
        if len(sorted_peaks) > 0:
            chart = LineChart()
            chart.title = "峰值振幅時間分布"
            chart.x_axis.title = "峰值排名"
            chart.y_axis.title = "振幅 (dB)"

            time_col = 2 if self.use_multi_band.get() else 1
            mag_col = time_col + 3

            data = Reference(ws, min_col=mag_col, min_row=1, max_row=len(sorted_peaks) + 1)
            chart.add_data(data, titles_from_data=True)

            ws.add_chart(chart, f"G2")

    def create_spectrum_sheet(self, wb):
        """創建整體頻譜資料工作表"""
        ws = wb.create_sheet("📈 頻譜資料")

        # 標題
        ws['A1'] = "頻率 (Hz)"
        ws['B1'] = "振幅 (dB)"

        # 設定標題樣式
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[cell].alignment = Alignment(horizontal="center")

        # 寫入資料（降採樣以減少檔案大小）
        step = max(1, len(self.xf_positive) // 1000)  # 最多1000個資料點

        for i in range(0, len(self.xf_positive), step):
            row = i // step + 2
            ws.cell(row=row, column=1, value=round(self.xf_positive[i], 2))
            ws.cell(row=row, column=2, value=round(self.yf_db[i], 2))

        # 調整欄寬
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15

    def create_timefreq_sheet(self, wb):
        """創建時頻分析資料工作表"""
        ws = wb.create_sheet("🌊 時頻資料")

        # 說明
        ws['A1'] = "時頻分析資料矩陣"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws.merge_cells('A1:D1')

        ws['A2'] = "說明：此工作表包含時間-頻率-振幅的三維資料"
        ws['A3'] = f"時間範圍: 0 - {self.stft_times[-1]:.2f} 秒"
        ws['A4'] = f"頻率範圍: 0 - {self.stft_frequencies[-1]:.0f} Hz"
        ws['A5'] = f"資料點數: {len(self.stft_times)} 個時間點 × {len(self.stft_frequencies)} 個頻率點"

        # 由於時頻資料過大，只提供摘要統計
        ws['A7'] = "時間點摘要統計"
        ws['A7'].font = Font(bold=True, color="FFFFFF")
        ws['A7'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A7:D7')

        ws['A8'] = "時間 (秒)"
        ws['B8'] = "最大振幅 (dB)"
        ws['C8'] = "平均振幅 (dB)"
        ws['D8'] = "峰值頻率 (Hz)"

        for cell in ['A8', 'B8', 'C8', 'D8']:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # 降採樣時間點（每隔N個取一個）
        step = max(1, len(self.stft_times) // 100)

        for i in range(0, len(self.stft_times), step):
            row = i // step + 9
            time_mag = self.stft_magnitude[:, i]

            ws.cell(row=row, column=1, value=round(self.stft_times[i], 2))
            ws.cell(row=row, column=2, value=round(np.max(time_mag), 2))
            ws.cell(row=row, column=3, value=round(np.mean(time_mag), 2))
            ws.cell(row=row, column=4, value=round(self.stft_frequencies[np.argmax(time_mag)], 1))

        # 調整欄寬
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 18

    def create_screenshots_sheet(self, wb):
        """創建截圖資訊工作表"""
        ws = wb.create_sheet("📸 截圖資訊")

        # 標題
        headers = ["排名", "時間 (秒)", "頻率 (Hz)", "振幅 (dB)", "檔案名稱"]

        if self.use_multi_band.get():
            headers.insert(1, "頻帶")

        if self.enable_audio_extract.get():
            headers.append("音訊檔案")

        # 寫入標題
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 寫入資料
        for screenshot in self.captured_screenshots:
            row = screenshot['rank'] + 1
            col = 1

            ws.cell(row=row, column=col, value=screenshot['rank'])
            col += 1

            if self.use_multi_band.get():
                ws.cell(row=row, column=col, value=screenshot.get('band_name', ''))
                col += 1

            ws.cell(row=row, column=col, value=round(screenshot['time'], 2))
            col += 1
            ws.cell(row=row, column=col, value=round(screenshot['frequency'], 1))
            col += 1
            ws.cell(row=row, column=col, value=round(screenshot['magnitude'], 2))
            col += 1
            ws.cell(row=row, column=col, value=screenshot['filepath'].name)
            col += 1

            if self.enable_audio_extract.get() and screenshot.get('audio_filepath'):
                ws.cell(row=row, column=col, value=screenshot['audio_filepath'].name)

        # 自動調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_multiband_sheet(self, wb):
        """創建多頻帶分析工作表"""
        ws = wb.create_sheet("🎼 多頻帶分析")

        # 標題
        ws['A1'] = "多頻帶分析結果"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws.merge_cells('A1:E1')

        row = 3

        # 各頻帶的統計資訊
        for band_name, band_data in self.multi_band_peaks.items():
            # 頻帶標題
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = f"頻帶: {band_name}"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            row += 1

            # 頻帶資訊
            freq_range = band_data['freq_range']
            ws[f'A{row}'] = "頻率範圍:"
            ws[f'B{row}'] = f"{freq_range[0]} - {freq_range[1]} Hz"
            row += 1

            peaks = band_data['peaks']
            ws[f'A{row}'] = "檢測峰值數:"
            ws[f'B{row}'] = len(peaks)
            row += 1

            if peaks:
                max_peak = band_data['max_peak']
                ws[f'A{row}'] = "最大峰值時間:"
                ws[f'B{row}'] = f"{max_peak['time']:.2f} 秒"
                row += 1

                ws[f'A{row}'] = "最大峰值振幅:"
                ws[f'B{row}'] = f"{max_peak['max_magnitude']:.2f} dB"
                row += 1

            row += 2

        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

    def cleanup(self):
        """清理資源"""
        def finish_cleanup():
            self.progress.config(mode='indeterminate')
            self.progress.stop()
            self.process_button.config(state="normal")
            self.is_processing = False

        # 刪除暫存音訊檔
        try:
            if os.path.exists(self.temp_audio_file):
                os.remove(self.temp_audio_file)
        except:
            pass

        self.root.after(0, finish_cleanup)

def main():
    """主函數"""
    root = tk.Tk()
    app = VideoAudioAnalyzer(root)

    # 設置視窗圖示（如果有的話）
    try:
        root.iconbitmap('icon.ico')  # 可選：添加應用程式圖示
    except:
        pass

    # 設置關閉事件
    def on_closing():
        if app.is_processing:
            if messagebox.askokcancel("退出", "分析正在進行中，確定要退出嗎？"):
                root.destroy()
        else:
            root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)

    # 啟動GUI
    root.mainloop()

if __name__ == "__main__":
    main()
