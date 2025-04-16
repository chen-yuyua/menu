import win32com.client
import os
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
import tkinter as tk
from tkinter import messagebox
import traceback
import sys

# 設置控制台輸出編碼為 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# 連接到Outlook
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

# 獲取收件箱和指定文件夾
inbox = outlook.GetDefaultFolder(6)  # 6 表示收件箱(受信トレイ)

# 更安全的文件夾存取方法
japan_folder = None
folder_name = "■BILの業務依頼"  # 新的資料夾名稱

try:
    japan_folder = inbox.Folders[folder_name]
except Exception as e:
    print(f"無法訪問文件夾 '{folder_name}': {str(e)}")
    print("可用的文件夾列表:")
    for folder in inbox.Folders:
        print(f" - {folder.Name}")
    sys.exit(1)

# 設置目標文件夾路徑
target_folder = r"D:\業務依賴信件"

# 設定日期範圍
end_date = datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999)
start_date = (end_date - timedelta(days=2)).replace(hour=0, minute=0, second=0, microsecond=0)

# 定義一個函數來移除文件名中的無效字符
def sanitize_filename(filename):
    return re.sub(r'[\\/*?:"<>|]', "", filename)

# 獲取並過濾郵件
def get_filtered_messages(folder):
    messages = folder.Items
    messages.Sort("[ReceivedTime]", True)
    filtered = []
    for msg in messages:
        try:
            if (msg.Class == 43 and  # 43 表示郵件項目
                start_date <= msg.ReceivedTime.replace(tzinfo=None) <= end_date and
                "業務依頼" in msg.Subject):
                filtered.append(msg)
        except AttributeError:
            # 如果項目沒有 ReceivedTime 或 Subject，跳過這個項目
            continue
    return filtered

# 獲取所有符合條件的郵件
# 1. 從收件匣(受信トレイ)本身獲取郵件
inbox_messages = get_filtered_messages(inbox)
# 2. 從BIL的業務依頼資料夾獲取郵件
japan_messages = get_filtered_messages(japan_folder)
# 合併所有郵件
all_messages = inbox_messages + japan_messages

try:
    print("開始處理Excel文件...")
    # 設置Excel文件路徑
    excel_file = r"D:\業務依賴信件\業務依頼メールまとめ.xlsx"

    # 創建或加載Excel文件
    wb = openpyxl.load_workbook(excel_file) if os.path.exists(excel_file) else openpyxl.Workbook()
    ws = wb.active

    print("設置表頭...")
    # 設置表頭
    headers = ["No.", "日付", "件名", "送信者"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="BIZ UDPゴシック", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    print("清空工作表...")
    # 清空工作表（保留表頭）
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 重置 last_row
    last_row = 1

    print("開始記錄郵件信息...")
    print(f"從收件匣(受信トレイ)獲取的郵件數: {len(inbox_messages)}")
    print(f"從{folder_name}資料夾獲取的郵件數: {len(japan_messages)}")
    print(f"總共處理的郵件數: {len(all_messages)}")

    # 保存符合條件的郵件並記錄到Excel
    for message in all_messages:
        received_time = message.ReceivedTime.replace(tzinfo=None)
        subject = str(message.Subject)

        # 創建日期文件夾
        date_folder = os.path.join(target_folder, received_time.strftime("%Y-%m-%d"))
        os.makedirs(date_folder, exist_ok=True)

        # 保存郵件
        file_name = f"{sanitize_filename(subject)}.msg"
        file_path = os.path.join(date_folder, file_name)
        if not os.path.exists(file_path):
            try:
                message.SaveAs(file_path)
                print(f"Saved email: {subject}")
            except Exception as e:
                print(f"Failed to save email: {subject}, error: {e}")

        # 記錄郵件信息到Excel
        last_row += 1
        for col, value in enumerate([last_row - 1, received_time.strftime("%Y/%m/%d"), subject, message.SenderName], start=1):
            cell = ws.cell(row=last_row, column=col, value=value)
            cell.font = Font(name="BIZ UDPゴシック")
            if col in [1, 2]:  # A和B列置中對齊
                cell.alignment = Alignment(horizontal='center', vertical='center')

    print("調整列寬...")
    # 調整列寬
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    print("保存Excel文件...")
    # 保存Excel文件
    wb.save(excel_file)

    print("顯示完成消息...")
    # 顯示完成消息
    root = tk.Tk()
    root.withdraw()  # 隱藏主窗口
    messagebox.showinfo("完成", f"信件資訊轉記成功，共處理 {len(all_messages)} 封郵件")

    print("程序完成")

except Exception as e:
    print(f"發生錯誤: {str(e)}")
    print("錯誤詳情:")
    print(traceback.format_exc())

print("Emails saved successfully and recorded in Excel")
input("按 Enter 鍵結束程式...")
