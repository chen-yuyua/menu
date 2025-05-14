import tkinter as tk
from tkinter import ttk, messagebox, font as tkfont
from tkcalendar import DateEntry
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta
import warnings

# 禁用 openpyxl 的警告
warnings.simplefilter(action='ignore', category=UserWarning)

# 定義無印良品風格的顏色
MUJI_COLORS = {
    "bg": "#F7F6F2",           # 淺米色背景
    "button": "#E9E4DD",       # 淺褐色按鈕
    "button_active": "#E0DAD0", # 按鈕點擊顏色
    "text": "#545454",         # 深灰色文字
    "accent": "#A89F91",       # 輕微強調色
    "highlight": "#D6CFC7",    # 高亮顏色
    "warning": "#E38B95",      # 警告色彩 (柔和紅)
    "success": "#708EB3",      # 成功色彩 (柔和藍，原為綠色)
    "border": "#DFDCD7",       # 邊框色彩
    "header": "#908C85"        # 表頭色彩
}

# 定義無印良品風格的字體
MUJI_FONT = {
    "family": "Noto Sans CJK JP",  # 首選字體
    "family_alt": "BIZ UDPゴシック", # 替代字體
    "size_small": 10,
    "size_normal": 12,
    "size_large": 14,
    "size_title": 18
}

class CountdownTimer:
    def __init__(self, label, duration, root):
        self.label = label
        self.duration = duration
        self.remaining = duration
        self.root = root
        self.update_label()

    def update_label(self):
        self.label.config(text=f"關閉倒數: {self.remaining}秒", fg=MUJI_COLORS["text"])
        if self.remaining > 0:
            self.remaining -= 1
            self.label.after(1000, self.update_label)
        else:
            self.label.config(text="關閉倒數: 0秒")
            self.root.destroy()

def read_excel_data(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook["2025年度"]
    data = []
    current_date = datetime.today().date()
    for row_idx, row in enumerate(sheet.iter_rows(min_row=12, max_row=1000, min_col=1, max_col=14), start=1):
        if isinstance(row[9].value, datetime) and row[13].value is None:
            delivery_date = row[9].value.date()
            if current_date <= delivery_date <= current_date + timedelta(days=7):
                values = [cell.value for cell in row]
                if all(values[i] not in (None, "") for i in [0, 6, 11]):
                    group = assign_group(values[11])
                    data.append([row_idx, values[0], values[6], delivery_date, values[11], group])
    workbook.close()
    data.sort(key=lambda x: abs((current_date - x[3]).days))
    return data

def read_specific_cells(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook["2025年度"]
    specific_cells = {
        "M1": sheet["M1"].value,
        "M2": sheet["M2"].value,
        "M3": sheet["M3"].value,
        "M4": sheet["M4"].value,
        "M5": sheet["M5"].value,
        "M6": sheet["M6"].value,
        "O1": sheet["O1"].value,
        "O2": sheet["O2"].value,
        "O3": sheet["O3"].value,
        "O4": sheet["O4"].value,
        "O5": sheet["O5"].value,
        "O6": sheet["O6"].value,
        "O7": sheet["O7"].value,
    }
    workbook.close()
    return specific_cells

def assign_group(name):
    group_a = ["呉汶珊", "陳雅瑄", "陳沅郁", "李恩柔", "翁紹綺", "陳岳揚"]
    if name in group_a:
        return "メカ2"
    else:
        return "メカ1"

def setup_muji_style():
    # 創建MUJI風格的ttk樣式
    style = ttk.Style()

    # 配置Scrollbar風格
    style.configure("TScrollbar",
                    background=MUJI_COLORS["bg"],
                    troughcolor=MUJI_COLORS["bg"],
                    arrowcolor=MUJI_COLORS["text"])

    # 配置分隔線風格
    style.configure("TSeparator",
                   background=MUJI_COLORS["accent"])

    # 配置表格風格
    style.configure("Treeview",
                   background=MUJI_COLORS["bg"],
                   fieldbackground=MUJI_COLORS["bg"],
                   foreground=MUJI_COLORS["text"])

    style.configure("Treeview.Heading",
                   background=MUJI_COLORS["header"],
                   foreground=MUJI_COLORS["bg"],
                   relief="flat")

    style.map("Treeview.Heading",
             background=[('pressed', MUJI_COLORS["accent"]),
                         ('active', MUJI_COLORS["accent"])])

def run_program_1():
    def update_info_label():
        file_path = r"\\Apbitwsh02\public\Project\TDD2\設計業務依頼\2025年度設計業務依頼台帳.xlsx"
        info_data = read_excel_data(file_path)
        specific_cells = read_specific_cells(file_path)
        current_date = datetime.today().strftime("%Y/%m/%d")

        for widget in canvas_frame.winfo_children():
            widget.destroy()

        # 靜態資訊欄
        static_info_frame = tk.Frame(canvas_frame, bg=MUJI_COLORS["bg"])
        static_info_frame.grid(row=0, column=0, columnspan=6, pady=10, sticky="nsew")

        # 計算總件數
        total_count = 0
        for key, value in specific_cells.items():
            if isinstance(value, (int, float)):
                total_count += value

        static_info_text = [
            ["■ボール業務件数：", "", "", "", f"總件數:{total_count}"],
            ["メカ1", "", "メカ2", "", "メカ∞"],
            [f"許映儂:{specific_cells['M1']}", "", f"陳雅瑄:{specific_cells['O3']}", "", f"陳俞源:{specific_cells['O1']}"],
            [f"葉羿廷:{specific_cells['M2']}", "", f"呉汶珊:{specific_cells['O2']}"],
            [f"何佳欣:{specific_cells['M3']}", "", f"陳沅郁:{specific_cells['O4']}"],
            [f"黄郁芸:{specific_cells['M4']}", "", f"李恩柔:{specific_cells['O5']}"],
            [f"王文豪:{specific_cells['M5']}", "", f"翁紹綺:{specific_cells['O6']}"],
            [f"林宜增:{specific_cells['M6']}", "", f"陳岳揚:{specific_cells['O7']}"],
        ]

        for i, texts in enumerate(static_info_text):
            for j, text in enumerate(texts):
                if text == "■ボール業務件数：":
                    label = tk.Label(static_info_frame, text=text,
                                    font=(MUJI_FONT["family_alt"], MUJI_FONT["size_title"]),
                                    fg=MUJI_COLORS["text"], bg=MUJI_COLORS["bg"])
                elif text == "メカ1" or text == "メカ2" or text == "メカ∞":
                    label = tk.Label(static_info_frame, text=text,
                                    font=(MUJI_FONT["family_alt"], MUJI_FONT["size_large"]),
                                    fg=MUJI_COLORS["accent"], bg=MUJI_COLORS["bg"])
                elif text.startswith("總件數:"):
                    label = tk.Label(static_info_frame, text=text,
                                    font=(MUJI_FONT["family_alt"], MUJI_FONT["size_large"], "bold"),
                                    fg=MUJI_COLORS["warning"], bg=MUJI_COLORS["bg"])
                else:
                    label = tk.Label(static_info_frame, text=text,
                                    font=(MUJI_FONT["family_alt"], MUJI_FONT["size_normal"]),
                                    fg=MUJI_COLORS["text"], bg=MUJI_COLORS["bg"])
                label.grid(row=i, column=j, padx=20, sticky="w")

        # 當前日期
        date_label = tk.Label(root, text=current_date,
                             font=(MUJI_FONT["family_alt"], MUJI_FONT["size_title"]),
                             fg=MUJI_COLORS["text"], bg=MUJI_COLORS["bg"])
        date_label.place(relx=1.0, rely=0, anchor='ne')

        # 倒數計時標籤
        countdown_var = tk.StringVar()
        countdown_label = tk.Label(root, textvariable=countdown_var,
                                  font=(MUJI_FONT["family_alt"], MUJI_FONT["size_small"]),
                                  fg=MUJI_COLORS["text"], bg=MUJI_COLORS["bg"])
        countdown_label.place(relx=1.0, rely=0.08, anchor='ne')

        # 更新倒數計時
        def update_countdown(seconds_left):
            if seconds_left > 0:
                countdown_var.set(f"視窗關閉倒數: {seconds_left}")
                root.after(1000, update_countdown, seconds_left - 1)
            else:
                root.destroy()

        # 開始倒數計時20秒
        update_countdown(20)

        # 分隔線
        separator = tk.Frame(canvas_frame, height=2, bg=MUJI_COLORS["accent"], width=500)
        separator.grid(row=len(static_info_text), column=0, columnspan=6, pady=10, sticky="ew")

        # 7日內通知書
        notice_label = tk.Label(canvas_frame, text="■7日內通知書",
                               font=(MUJI_FONT["family_alt"], MUJI_FONT["size_title"]),
                               fg=MUJI_COLORS["text"], bg=MUJI_COLORS["bg"])
        notice_label.grid(row=len(static_info_text) + 1, column=0, columnspan=6, pady=10, sticky="w")

        # 表格背景框架
        table_frame = tk.Frame(canvas_frame, bg=MUJI_COLORS["bg"], padx=20)
        table_frame.grid(row=len(static_info_text) + 2, column=0, columnspan=6, sticky="nsew")

        # 表頭
        headers = ["No.", "台帳", "內容", "納期", "擔當者", "グループ"]
        for i, header in enumerate(headers):
            header_label = tk.Label(table_frame, text=header,
                                   font=(MUJI_FONT["family_alt"], MUJI_FONT["size_normal"], "bold"),
                                   anchor="w", bg=MUJI_COLORS["header"], fg=MUJI_COLORS["bg"],
                                   padx=5, pady=5)
            header_label.grid(row=0, column=i, sticky="nsew")

        for i, row in enumerate(info_data, start=1):
            for j, value in enumerate(row):
                font = (MUJI_FONT["family_alt"], MUJI_FONT["size_normal"])
                anchor = "w"
                font_color = MUJI_COLORS["text"]
                bg_color = MUJI_COLORS["bg"]

                if j == 0:
                    value = i
                if j == 3 and value == datetime.today().date():
                    font_color = MUJI_COLORS["warning"]
                    # 納期欄位(當天)使用背景色突顯
                    bg_color = "#FFEBEE"  # 非常柔和的淺紅色背景
                elif j == 4 and value in ["呉汶珊", "陳雅瑄", "陳沅郁", "李恩柔", "翁紹綺", "陳岳揚"]:
                    font_color = MUJI_COLORS["success"]
                    font = (MUJI_FONT["family_alt"], MUJI_FONT["size_normal"], "bold")
                elif j == 5 and value == "メカ2":
                    font_color = MUJI_COLORS["success"]

                data_label = tk.Label(table_frame, text=value, font=font, anchor=anchor,
                                     fg=font_color, bg=bg_color, padx=5, pady=5)
                if j == 2:
                    data_label.configure(wraplength=700)
                data_label.grid(row=i, column=j, sticky="nsew")

        for child in table_frame.winfo_children():
            child.grid_configure(padx=2, pady=2, sticky="nsew")
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(len(info_data) + 1, weight=1)

        canvas.configure(scrollregion=canvas.bbox("all"))

        # 調整視窗大小
        canvas.update_idletasks()
        frame_width = canvas_frame.winfo_reqwidth()
        frame_height = canvas_frame.winfo_reqheight()
        root.geometry(f"{frame_width+20}x{frame_height+20}")

    root = tk.Toplevel()
    root.title("7日內通知書掲示板")
    root.configure(bg=MUJI_COLORS["bg"])

    # 設置MUJI風格
    setup_muji_style()

    canvas = tk.Canvas(root, bg=MUJI_COLORS["bg"], highlightthickness=0)
    canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)

    scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    canvas_frame = tk.Frame(canvas, bg=MUJI_COLORS["bg"])
    canvas.create_window((0, 0), window=canvas_frame, anchor="nw")

    canvas_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.configure(yscrollcommand=scrollbar.set)

    update_info_label()
    root.mainloop()

def program_1():
    def read_excel_data(file_path):
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook["2025年 未発行図面"]
        data = []
        current_date = datetime.today().date()
        end_date = current_date + timedelta(days=30)

        for row in sheet.iter_rows(min_row=3):
            data_date = row[12].value
            check_column_l = row[11].value  # L 列
            check_column_n = row[13].value
            if (isinstance(data_date, datetime) and
                (check_column_l is None or check_column_l == "") and
                (check_column_n is None or check_column_n == "") and
                current_date <= data_date.date() <= end_date):
                row_data = [row[0].value, row[0].value, row[1].value, row[3].value, row[5].value, row[6].value, row[8].value, data_date.date()]
                data.append(row_data)

        workbook.close()
        data.sort(key=lambda x: abs((current_date - x[7]).days))
        return data

    def update_countdown(root, countdown_label, seconds_left):
        if seconds_left > 0:
            countdown_label.config(text=f"關閉倒數計時: {seconds_left}")
            root.after(1000, update_countdown, root, countdown_label, seconds_left - 1)
        else:
            root.destroy()

    file_path = r"\\Apngo-peach\public\Dev_Share\【進捗】仕様新設進捗管理\2025仕様新設未発行図面【自動転記トライアルVer.】.xlsm"
    excel_data = read_excel_data(file_path)

    root = tk.Toplevel()
    root.title("データ予定日")
    root.configure(bg=MUJI_COLORS["bg"])

    # 設置MUJI風格
    setup_muji_style()

    main_frame = tk.Frame(root, padx=20, pady=20, bg=MUJI_COLORS["bg"])
    main_frame.pack(fill="both", expand=True)

    header_frame = tk.Frame(main_frame, bg=MUJI_COLORS["bg"])
    header_frame.grid(row=0, column=0, columnspan=9, pady=10, sticky="ew")

    title_label = tk.Label(header_frame, text="■30日内データ予定",
                          font=(MUJI_FONT["family_alt"], MUJI_FONT["size_title"]),
                          fg=MUJI_COLORS["text"], bg=MUJI_COLORS["bg"])
    title_label.pack(side="left")

    info_frame = tk.Frame(header_frame, bg=MUJI_COLORS["bg"])
    info_frame.pack(side="right", anchor="ne")

    info_label = tk.Label(info_frame, text="7日内は赤字表示",
                         font=(MUJI_FONT["family_alt"], MUJI_FONT["size_normal"]),
                         fg=MUJI_COLORS["warning"], bg=MUJI_COLORS["bg"])
    info_label.pack(anchor='e')

    countdown = 20
    countdown_label = tk.Label(info_frame, text=f"關閉倒數計時: {countdown}",
                              font=(MUJI_FONT["family_alt"], MUJI_FONT["size_small"]),
                              fg=MUJI_COLORS["text"], bg=MUJI_COLORS["bg"])
    countdown_label.pack(anchor='e')

    separator1 = ttk.Separator(main_frame, orient="horizontal")
    separator1.grid(row=1, column=0, columnspan=9, sticky="ew", pady=(0, 10))

    headers = ["No.", "機種", "仕様", "生産月", "部品コード", "品名", "担当者", "データ予定日"]
    for col, header in enumerate(headers):
        header_label = tk.Label(main_frame, text=header,
                               font=(MUJI_FONT["family_alt"], MUJI_FONT["size_normal"], "bold"),
                               padx=10, pady=5, bg=MUJI_COLORS["header"], fg=MUJI_COLORS["bg"])
        header_label.grid(row=2, column=col, sticky="nsew")

    for col in range(len(headers)):
        main_frame.grid_columnconfigure(col, weight=1)

    separator2 = ttk.Separator(main_frame, orient="horizontal")
    separator2.grid(row=3, column=0, columnspan=9, sticky="ew", pady=(10, 0))

    for row_idx, row_data in enumerate(excel_data, start=1):
        row_data[0] = row_idx
        for col_idx, cell_value in enumerate(row_data):
            font_color = MUJI_COLORS["text"]
            bg_color = MUJI_COLORS["bg"]

            if col_idx == 7:
                days_diff = (cell_value - datetime.today().date()).days
                if days_diff < 0:
                    bg_color = "#FFCDD2"  # 更明顯但仍柔和的紅色背景
                    font_color = "#B71C1C"  # 深紅色文字
                elif 0 <= days_diff <= 7:
                    font_color = MUJI_COLORS["warning"]
                    bg_color = "#FFEBEE"  # 非常柔和的淺紅色背景

            data_label = tk.Label(main_frame, text=cell_value,
                                 font=(MUJI_FONT["family_alt"], MUJI_FONT["size_normal"]),
                                 padx=10, pady=5, fg=font_color, bg=bg_color)
            data_label.grid(row=row_idx+3, column=col_idx, sticky="nsew")

    root.after(1000, update_countdown, root, countdown_label, countdown)

    root.update_idletasks()
    frame_width = root.winfo_reqwidth() + 40
    frame_height = root.winfo_reqheight() + 40
    root.geometry(f"{frame_width}x{frame_height}")

    root.mainloop()

def program_2():
    def read_excel_data(file_path):
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook["2025年 未発行図面"]
        data = []
        current_date = datetime.today()
        start_date = current_date - timedelta(days=7)
        end_date = current_date + timedelta(days=14)

        for row in sheet.iter_rows(min_row=3):
            if row[11].value in (None, ""):  # L 列為空格
                data_date = row[9].value  # J 列的日期
                if isinstance(data_date, datetime) and start_date.date() <= data_date.date() <= end_date.date():
                    row_data = [None, row[0].value, row[1].value, row[3].value, row[5].value, row[6].value, row[8].value, data_date.date()]
                    data.append(row_data)

        workbook.close()
        data.sort(key=lambda x: x[7])  # 按照日期排序
        return data

    def update_countdown(root, countdown_label, seconds_left):
        if seconds_left > 0:
            countdown_label.config(text=f"關閉倒數計時: {seconds_left}")
            root.after(1000, update_countdown, root, countdown_label, seconds_left - 1)
        else:
            root.destroy()

    file_path = r"\\Apngo-peach\public\Dev_Share\【進捗】仕様新設進捗管理\2025仕様新設未発行図面【自動転記トライアルVer.】.xlsm"
    excel_data = read_excel_data(file_path)

    root = tk.Toplevel()
    root.title("未発行図面")
    root.configure(bg=MUJI_COLORS["bg"])

    # 設置MUJI風格
    setup_muji_style()

    main_frame = tk.Frame(root, padx=20, pady=20, bg=MUJI_COLORS["bg"])
    main_frame.pack(fill="both", expand=True)

    header_frame = tk.Frame(main_frame, bg=MUJI_COLORS["bg"])
    header_frame.grid(row=0, column=0, columnspan=8, pady=10, sticky="ew")

    title_label = tk.Label(header_frame, text="■14日内未発行図面",
                          font=(MUJI_FONT["family_alt"], MUJI_FONT["size_title"]),
                          fg=MUJI_COLORS["text"], bg=MUJI_COLORS["bg"])
    title_label.pack(side="left")

    info_frame = tk.Frame(header_frame, bg=MUJI_COLORS["bg"])
    info_frame.pack(side="right", anchor="ne")

    info_label = tk.Label(info_frame, text="7日内は赤字",
                         font=(MUJI_FONT["family_alt"], MUJI_FONT["size_normal"]),
                         fg=MUJI_COLORS["warning"], bg=MUJI_COLORS["bg"])
    info_label.pack(anchor='e')

    countdown = 20
    countdown_label = tk.Label(info_frame, text=f"關閉倒數計時: {countdown}",
                              font=(MUJI_FONT["family_alt"], MUJI_FONT["size_small"]),
                              fg=MUJI_COLORS["text"], bg=MUJI_COLORS["bg"])
    countdown_label.pack(anchor='e')

    separator1 = ttk.Separator(main_frame, orient="horizontal")
    separator1.grid(row=1, column=0, columnspan=8, sticky="ew", pady=(0, 10))

    headers = ["No.", "機種", "仕様", "生産月", "部品コード", "品名", "担当者", "納期"]
    for col, header in enumerate(headers):
        header_label = tk.Label(main_frame, text=header,
                               font=(MUJI_FONT["family_alt"], MUJI_FONT["size_normal"], "bold"),
                               padx=10, pady=5, bg=MUJI_COLORS["header"], fg=MUJI_COLORS["bg"])
        header_label.grid(row=2, column=col, sticky="nsew")

    for col in range(len(headers)):
        main_frame.grid_columnconfigure(col, weight=1)

    separator2 = ttk.Separator(main_frame, orient="horizontal")
    separator2.grid(row=3, column=0, columnspan=8, sticky="ew", pady=(10, 0))

    for row_idx, row_data in enumerate(excel_data, start=1):
        row_data[0] = row_idx
        for col_idx, cell_value in enumerate(row_data):
            font_color = MUJI_COLORS["text"]
            bg_color = MUJI_COLORS["bg"]

            if col_idx == 7:
                days_diff = (cell_value - datetime.today().date()).days
                if days_diff < 0:
                    bg_color = "#FFCDD2"  # 更明顯但仍柔和的紅色背景
                    font_color = "#B71C1C"  # 深紅色文字
                elif 0 <= days_diff <= 7:
                    font_color = MUJI_COLORS["warning"]
                    bg_color = "#FFEBEE"  # 非常柔和的淺紅色背景

            data_label = tk.Label(main_frame, text=cell_value,
                                 font=(MUJI_FONT["family_alt"], MUJI_FONT["size_normal"]),
                                 padx=10, pady=5, fg=font_color, bg=bg_color)
            data_label.grid(row=row_idx+3, column=col_idx, sticky="nsew")

    root.after(1000, update_countdown, root, countdown_label, countdown)

    root.update_idletasks()
    frame_width = root.winfo_reqwidth() + 40
    frame_height = root.winfo_reqheight() + 40
    root.geometry(f"{frame_width}x{frame_height}")

    root.mainloop()

def create_rounded_button(parent, text, command, width=15):
    frame = tk.Frame(parent, bg=MUJI_COLORS["bg"])

    button = tk.Button(
        frame,
        text=text,
        command=command,
        font=(MUJI_FONT["family_alt"], MUJI_FONT["size_normal"]),
        bg=MUJI_COLORS["button"],
        fg=MUJI_COLORS["text"],
        relief="flat",
        width=width,
        padx=15,
        pady=8,
        borderwidth=0,
        activebackground=MUJI_COLORS["button_active"],
        activeforeground=MUJI_COLORS["text"]
    )
    button.pack(padx=10, pady=10)

    # 按鈕懸停效果
    def on_enter(e):
        button['background'] = MUJI_COLORS["button_active"]

    def on_leave(e):
        button['background'] = MUJI_COLORS["button"]

    button.bind("<Enter>", on_enter)
    button.bind("<Leave>", on_leave)

    return frame

# 主程序
main_root = tk.Tk()
main_root.title("分室業務関連")
main_root.configure(bg=MUJI_COLORS["bg"])

# 標題框架
title_frame = tk.Frame(main_root, bg=MUJI_COLORS["bg"], pady=15)
title_frame.pack(fill="x")

title_label = tk.Label(
    title_frame,
    text="分室業務関連",
    font=(MUJI_FONT["family_alt"], MUJI_FONT["size_title"]),
    fg=MUJI_COLORS["text"],
    bg=MUJI_COLORS["bg"]
)
title_label.pack()

# 按鈕框架
button_frame = tk.Frame(main_root, bg=MUJI_COLORS["bg"], padx=20, pady=10)
button_frame.pack(fill="both", expand=True)

# 創建圓角按鈕
button1_frame = create_rounded_button(button_frame, "7日間通知書", run_program_1)
button1_frame.pack(pady=8)

button2_frame = create_rounded_button(button_frame, "データ予定日早見表", program_1)
button2_frame.pack(pady=8)

button3_frame = create_rounded_button(button_frame, "未発行図面", program_2)
button3_frame.pack(pady=8)

# 底部框架
footer_frame = tk.Frame(main_root, bg=MUJI_COLORS["bg"], pady=15)
footer_frame.pack(fill="x")

# 版權資訊
copyright_label = tk.Label(
    footer_frame,
    text="© 2025 分室業務関連 Ver 1.2\n作成者：陳兪源",
    font=(MUJI_FONT["family_alt"], MUJI_FONT["size_small"]),
    fg=MUJI_COLORS["accent"],
    bg=MUJI_COLORS["bg"]
)
copyright_label.pack()

# 調整主視窗大小和位置
main_root.update_idletasks()
frame_width = max(button_frame.winfo_reqwidth(), 300)
frame_height = title_frame.winfo_reqheight() + button_frame.winfo_reqheight() + footer_frame.winfo_reqheight() + 60
main_root.geometry(f"{frame_width}x{frame_height}")

# 置中視窗
screen_width = main_root.winfo_screenwidth()
screen_height = main_root.winfo_screenheight()
x = (screen_width - frame_width) // 2
y = (screen_height - frame_height) // 2
main_root.geometry(f"+{x}+{y}")

main_root.mainloop()
