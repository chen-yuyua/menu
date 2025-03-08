import tkinter as tk
from tkinter import ttk, messagebox, font as tkfont
from tkcalendar import DateEntry
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta
import warnings

# 禁用 openpyxl 的警告
warnings.simplefilter(action='ignore', category=UserWarning)

class CountdownTimer:
    def __init__(self, label, duration, root):
        self.label = label
        self.duration = duration
        self.remaining = duration
        self.root = root
        self.update_label()

    def update_label(self):
        self.label.config(text=f"關閉倒數: {self.remaining}秒")
        if self.remaining > 0:
            self.remaining -= 1
            self.label.after(1000, self.update_label)
        else:
            self.label.config(text="關閉倒數: 0秒")
            self.root.destroy()

def read_excel_data(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook["2024年度"]
    data = []
    current_date = datetime.today().date()
    for row_idx, row in enumerate(sheet.iter_rows(min_row=12, max_row=1000, min_col=1, max_col=14), start=1):
        if isinstance(row[9].value, datetime) and row[13].value is None:
            delivery_date = row[9].value.date()
            if current_date <= delivery_date <= current_date + timedelta(days=7):
                values = [cell.value for cell in row]
                if all(values[i] not in (None, "") for i in [0, 6, 11]):
                    group = assign_group(values[11])
                    data.append([row_idx, values[0], values[6], delivery_date, values[11], group])
    workbook.close()
    data.sort(key=lambda x: abs((current_date - x[3]).days))
    return data

def read_specific_cells(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook["2024年度"]
    specific_cells = {
        "M1": sheet["M1"].value,
        "M2": sheet["M2"].value,
        "M3": sheet["M3"].value,
        "M4": sheet["M4"].value,
        "M5": sheet["M5"].value,
        "M6": sheet["M6"].value,
        "O1": sheet["O1"].value,
        "O2": sheet["O2"].value,
        "O3": sheet["O3"].value,
        "O4": sheet["O4"].value,
        "O5": sheet["O5"].value,
        "O6": sheet["O6"].value,
        "O7": sheet["O7"].value,
    }
    workbook.close()
    return specific_cells

def assign_group(name):
    group_a = ["呉汶珊", "陳雅瑄", "陳沅郁", "李恩柔", "翁紹綺", "陳岳揚"]
    if name in group_a:
        return "メカ2"
    else:
        return "メカ1"

def run_program_1():
    def update_info_label():
        file_path = r"\\Apbitwsh02\public\Project\TDD2\設計業務依頼\2024年度設計業務依頼台帳.xlsx"
        info_data = read_excel_data(file_path)
        specific_cells = read_specific_cells(file_path)
        current_date = datetime.today().strftime("%Y/%m/%d")

        for widget in canvas_frame.winfo_children():
            widget.destroy()

        # 靜態資訊欄
        static_info_frame = tk.Frame(canvas_frame)
        static_info_frame.grid(row=0, column=0, columnspan=6, pady=10, sticky="nsew")

        static_info_text = [
            ["■ボール業務件数：", "", ""],
            ["メカ1", "", "メカ2", "", "メカ∞"],
            [f"許映儂:{specific_cells['M1']}", "", f"陳雅瑄:{specific_cells['O3']}", "", f"陳俞源:{specific_cells['O1']}"],
            [f"葉羿廷:{specific_cells['M2']}", "", f"呉汶珊:{specific_cells['O2']}"],
            [f"何佳欣:{specific_cells['M3']}", "", f"陳沅郁:{specific_cells['O4']}"],
            [f"黄郁芸:{specific_cells['M4']}", "", f"李恩柔:{specific_cells['O5']}"],
            [f"王文豪:{specific_cells['M5']}", "", f"翁紹綺:{specific_cells['O6']}"],
            [f"林宜增:{specific_cells['M6']}", "", f"陳岳揚:{specific_cells['O7']}"],
        ]

        for i, texts in enumerate(static_info_text):
            for j, text in enumerate(texts):
                if text == "■ボール業務件数：":
                    label = tk.Label(static_info_frame, text=text, font=("BIZ UDPゴシック", 20, "bold"))
                elif text == "メカ1" or text == "メカ2" or text == "メカ∞":
                    label = tk.Label(static_info_frame, text=text, font=("BIZ UDPゴシック", 14, "bold"))
                else:
                    label = tk.Label(static_info_frame, text=text, font=("BIZ UDPゴシック", 12))
                label.grid(row=i, column=j, padx=20, sticky="w")

        # 當前日期
        date_label = tk.Label(root, text=current_date, font=("BIZ UDPゴシック", 20, "bold"))
        date_label.place(relx=1.0, rely=0, anchor='ne')

        # 倒數計時標籤
        countdown_var = tk.StringVar()
        countdown_label = tk.Label(root, textvariable=countdown_var, font=("BIZ UDPゴシック", 10))
        countdown_label.place(relx=1.0, rely=0.08, anchor='ne')

        # 更新倒數計時
        def update_countdown(seconds_left):
            if seconds_left > 0:
                countdown_var.set(f"視窗關閉倒數: {seconds_left}")
                root.after(1000, update_countdown, seconds_left - 1)
            else:
                root.destroy()

        # 開始倒數計時20秒
        update_countdown(20)

        # 分隔線
        separator = tk.Frame(canvas_frame, height=5, bg="black", width=500)
        separator.grid(row=len(static_info_text), column=0, columnspan=6, pady=10, sticky="ew")

        # 7日內通知書
        notice_label = tk.Label(canvas_frame, text="■7日內通知書", font=("BIZ UDPゴシック", 20, "bold"))
        notice_label.grid(row=len(static_info_text) + 1, column=0, columnspan=6, pady=10, sticky="w")

        # 表格背景框架
        table_frame = tk.Frame(canvas_frame, bg="white", padx=20)
        table_frame.grid(row=len(static_info_text) + 2, column=0, columnspan=6, sticky="nsew")

        # 表頭
        headers = ["No.", "台帳", "內容", "納期", "擔當者", "グループ"]
        for i, header in enumerate(headers):
            header_label = tk.Label(table_frame, text=header, font=("BIZ UDPゴシック", 12, "bold"), anchor="w", bg="white")
            header_label.grid(row=0, column=i, sticky="nsew")

        for i, row in enumerate(info_data, start=1):
            for j, value in enumerate(row):
                font = ("BIZ UDPゴシック", 12)
                anchor = "w"
                font_color = "#000000"
                if j == 0:
                    value = i
                if j == 3 and value == datetime.today().date():
                    font_color = "#D20062"
                elif j == 4 and value in ["呉汶珊", "陳雅瑄", "陳沅郁", "李恩柔", "翁紹綺", "陳岳揚"]:
                    font_color = "#0072E3"
                    font = ("BIZ UDPゴシック", 12, "bold")
                elif j == 5 and value == "メカ2":
                    font_color = "#0072E3"

                data_label = tk.Label(table_frame, text=value, font=font, anchor=anchor, fg=font_color, bg="white")
                if j == 2:
                    data_label.configure(wraplength=700)
                data_label.grid(row=i, column=j, sticky="nsew")

        for child in table_frame.winfo_children():
            child.grid_configure(padx=5, pady=5, sticky="nsew")
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(len(info_data) + 1, weight=1)

        canvas.configure(scrollregion=canvas.bbox("all"))

        # 調整視窗大小
        canvas.update_idletasks()
        frame_width = canvas_frame.winfo_reqwidth()
        frame_height = canvas_frame.winfo_reqheight()
        root.geometry(f"{frame_width+20}x{frame_height+20}")

    root = tk.Toplevel()
    root.title("7日內通知書掲示板")

    canvas = tk.Canvas(root)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    canvas_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=canvas_frame, anchor="nw")

    canvas_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.configure(yscrollcommand=scrollbar.set)

    update_info_label()
    root.mainloop()

def program_1():
    def read_excel_data(file_path):
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook["2024年 未発行図面"]
        data = []
        current_date = datetime.today().date()
        end_date = current_date + timedelta(days=30)

        for row in sheet.iter_rows(min_row=3):
            data_date = row[12].value
            check_column_l = row[11].value  # L 列
            check_column_n = row[13].value
            if (isinstance(data_date, datetime) and
                (check_column_l is None or check_column_l == "") and
                (check_column_n is None or check_column_n == "") and
                current_date <= data_date.date() <= end_date):
                row_data = [row[0].value, row[0].value, row[1].value, row[3].value, row[5].value, row[6].value, row[8].value, data_date.date()]
                data.append(row_data)

        workbook.close()
        data.sort(key=lambda x: abs((current_date - x[7]).days))
        return data

    def update_countdown(root, countdown_label, seconds_left):
        if seconds_left > 0:
            countdown_label.config(text=f"關閉倒數計時: {seconds_left}")
            root.after(1000, update_countdown, root, countdown_label, seconds_left - 1)
        else:
            root.destroy()

    file_path = r"\\Apngo-peach\public\Dev_Share\【進捗】仕様新設進捗管理\2024仕様新設未発行図面.xlsx"
    excel_data = read_excel_data(file_path)

    root = tk.Toplevel()
    root.title("データ予定日")

    main_frame = tk.Frame(root, padx=20, pady=20)
    main_frame.pack(fill="both", expand=True)

    header_frame = tk.Frame(main_frame)
    header_frame.grid(row=0, column=0, columnspan=9, pady=10, sticky="ew")

    title_label = tk.Label(header_frame, text="■30日内データ予定", font=("BIZ UDPゴシック", 20, "bold"))
    title_label.pack(side="left")

    info_frame = tk.Frame(header_frame)
    info_frame.pack(side="right", anchor="ne")

    info_label = tk.Label(info_frame, text="7日内は赤字表示", font=("BIZ UDPゴシック", 15), fg="red")
    info_label.pack(anchor='e')

    countdown = 20
    countdown_label = tk.Label(info_frame, text=f"關閉倒數計時: {countdown}", font=("BIZ UDPゴシック", 10))
    countdown_label.pack(anchor='e')

    separator1 = ttk.Separator(main_frame, orient="horizontal")
    separator1.grid(row=1, column=0, columnspan=9, sticky="ew", pady=(0, 10))

    headers = ["No.", "機種", "仕様", "生産月", "部品コード", "品名", "担当者", "データ予定日"]
    for col, header in enumerate(headers):
        header_label = tk.Label(main_frame, text=header, font=("BIZ UDPゴシック", 12, "bold"), padx=10, pady=5)
        header_label.grid(row=2, column=col, sticky="nsew")

    for col in range(len(headers)):
        main_frame.grid_columnconfigure(col, weight=1)

    separator2 = ttk.Separator(main_frame, orient="horizontal")
    separator2.grid(row=3, column=0, columnspan=9, sticky="ew", pady=(10, 0))

    for row_idx, row_data in enumerate(excel_data, start=1):
        row_data[0] = row_idx
        for col_idx, cell_value in enumerate(row_data):
            font_color = "#000000"
            bg_color = "#FFFFFF"
            if col_idx == 7:
                days_diff = (cell_value - datetime.today().date()).days
                if days_diff < 0:
                    bg_color = "#FFCCCC"
                elif 0 <= days_diff <= 7:
                    font_color = "#FF0000"
            data_label = tk.Label(main_frame, text=cell_value, font=("BIZ UDPゴシック", 12), padx=10, pady=5, fg=font_color, bg=bg_color)
            data_label.grid(row=row_idx+3, column=col_idx, sticky="nsew")

    root.after(1000, update_countdown, root, countdown_label, countdown)

    root.update_idletasks()
    frame_width = root.winfo_reqwidth() + 40
    frame_height = root.winfo_reqheight() + 40
    root.geometry(f"{frame_width}x{frame_height}")

    root.mainloop()

def program_2():
    def read_excel_data(file_path):
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook["2024年 未発行図面"]
        data = []
        current_date = datetime.today()
        start_date = current_date - timedelta(days=7)
        end_date = current_date + timedelta(days=14)

        for row in sheet.iter_rows(min_row=3):
            if row[11].value in (None, ""):  # L 列為空格
                data_date = row[9].value  # J 列的日期
                if isinstance(data_date, datetime) and start_date.date() <= data_date.date() <= end_date.date():
                    row_data = [None, row[0].value, row[1].value, row[3].value, row[5].value, row[6].value, row[8].value, data_date.date()]
                    data.append(row_data)

        workbook.close()
        data.sort(key=lambda x: x[7])  # 按照日期排序
        return data

    def update_countdown(root, countdown_label, seconds_left):
        if seconds_left > 0:
            countdown_label.config(text=f"關閉倒數計時: {seconds_left}")
            root.after(1000, update_countdown, root, countdown_label, seconds_left - 1)
        else:
            root.destroy()

    file_path = r"\\Apngo-peach\public\Dev_Share\【進捗】仕様新設進捗管理\2024仕様新設未発行図面.xlsx"
    excel_data = read_excel_data(file_path)

    root = tk.Toplevel()
    root.title("未発行図面")

    main_frame = tk.Frame(root, padx=20, pady=20)
    main_frame.pack(fill="both", expand=True)

    header_frame = tk.Frame(main_frame)
    header_frame.grid(row=0, column=0, columnspan=8, pady=10, sticky="ew")

    title_label = tk.Label(header_frame, text="■14日内未発行図面", font=("BIZ UDPゴシック", 20, "bold"))
    title_label.pack(side="left")

    info_frame = tk.Frame(header_frame)
    info_frame.pack(side="right", anchor="ne")

    info_label = tk.Label(info_frame, text="7日内は赤字", font=("BIZ UDPゴシック", 15), fg="red")
    info_label.pack(anchor='e')

    countdown = 20
    countdown_label = tk.Label(info_frame, text=f"關閉倒數計時: {countdown}", font=("BIZ UDPゴシック", 10))
    countdown_label.pack(anchor='e')

    separator1 = ttk.Separator(main_frame, orient="horizontal")
    separator1.grid(row=1, column=0, columnspan=8, sticky="ew", pady=(0, 10))

    headers = ["No.", "機種", "仕様", "生産月", "部品コード", "品名", "担当者", "納期"]
    for col, header in enumerate(headers):
        header_label = tk.Label(main_frame, text=header, font=("BIZ UDPゴシック", 12, "bold"), padx=10, pady=5)
        header_label.grid(row=2, column=col, sticky="nsew")

    for col in range(len(headers)):
        main_frame.grid_columnconfigure(col, weight=1)

    separator2 = ttk.Separator(main_frame, orient="horizontal")
    separator2.grid(row=3, column=0, columnspan=8, sticky="ew", pady=(10, 0))

    for row_idx, row_data in enumerate(excel_data, start=1):
        row_data[0] = row_idx
        for col_idx, cell_value in enumerate(row_data):
            font_color = "#000000"
            bg_color = "#FFFFFF"
            if col_idx == 7:
                days_diff = (cell_value - datetime.today().date()).days
                if days_diff < 0:
                    bg_color = "#FFCCCC"
                elif 0 <= days_diff <= 7:
                    font_color = "#FF0000"
            data_label = tk.Label(main_frame, text=cell_value, font=("BIZ UDPゴシック", 12), padx=10, pady=5, fg=font_color, bg=bg_color)
            data_label.grid(row=row_idx+3, column=col_idx, sticky="nsew")

    root.after(1000, update_countdown, root, countdown_label, countdown)

    root.update_idletasks()
    frame_width = root.winfo_reqwidth() + 40
    frame_height = root.winfo_reqheight() + 40
    root.geometry(f"{frame_width}x{frame_height}")

    root.mainloop()

def program_3():
    def show_info():
        selected_name = combo_name.get()
        if not selected_name:
            messagebox.showerror("錯誤", "請選擇姓名")
            return

        try:
            # 打開Excel文件
            workbook = load_workbook(r'\\Apbitwsh02\public\Project\TDD2\設計業務依頼\2024年度設計業務依頼台帳.xlsx')
            sheet = workbook['2024年度']

            # 清除現有的Treeview數據
            for item in tree_info.get_children():
                tree_info.delete(item)

            # 遍歷從L12開始的列，查找符合條件的行
            data = extract_data(sheet, selected_name)
            update_treeview(data)

            if not data:
                messagebox.showinfo("結果", "未找到符合條件的記錄")
        except Exception as e:
            messagebox.showerror("錯誤", f"讀取文件時發生錯誤: {e}")

    def extract_data(sheet, selected_name):
        data = []
        for row in sheet.iter_rows(min_row=12, min_col=1, max_col=17):  # A to Q columns
            if row[11].value == selected_name and not row[15].value:  # L and P columns
                account_number = row[0].value
                content = row[6].value
                business_status = row[13].value if isinstance(row[13].value, datetime) else row[13].value
                remarks = row[16].value
                business_status = business_status.strftime('%Y/%m/%d') if isinstance(business_status, datetime) else business_status
                data.append((account_number, content, business_status, remarks))
        return data

    def update_treeview(data):
        for i, (account_number, content, business_status, remarks) in enumerate(data):
            tree_info.insert('', 'end', values=(i + 1, account_number, content, business_status, remarks))

        # 自動調整其他欄位寬度
        for col in tree_info['columns']:
            if col != "No.":
                max_width = tkfont.Font(font=font).measure(col)
                for item in tree_info.get_children():
                    cell_value = tree_info.set(item, col)
                    max_width = max(max_width, tkfont.Font(font=font).measure(cell_value))
                tree_info.column(col, width=max_width)

    def setup_ui(root):
        global font, combo_name, tree_info

        # 設置字體
        font = ("BIZ UDPゴシック", 12)

        # 姓名區域
        frame_name = tk.Frame(root)
        frame_name.pack(pady=10, padx=10, fill='x')

        label_name = tk.Label(frame_name, text="姓名:", font=font)
        label_name.pack(side='left')

        combo_name = ttk.Combobox(frame_name, font=font)
        combo_name.pack(side='left', padx=5)

        # 添加下拉式選單項目
        name_list = ["許映儂", "葉羿廷", "何佳欣", "黄郁芸", "王文豪", "林宜增", "陳雅瑄", "呉汶珊", "陳沅郁", "李恩柔", "翁紹綺", "陳岳揚", "陳俞源"]
        combo_name['values'] = name_list

        button_show_info = tk.Button(frame_name, text="資訊顯示", font=font, command=show_info)
        button_show_info.pack(side='left', padx=5)

        countdown_label = tk.Label(frame_name, text="", font=font, fg="red")
        countdown_label.pack(side='left', padx=5)

        # 初始化倒數計時器
        countdown_timer = CountdownTimer(countdown_label, 60, root)

        # 未檢收通知書資訊區域
        frame_info = tk.LabelFrame(root, text="未檢收通知書資訊", font=font)
        frame_info.pack(pady=10, padx=10, fill='both', expand=True)

        tree_info = ttk.Treeview(frame_info, columns=("No.", "台帳番號", "內容", "業務狀況", "備考"), show='headings', height=10)
        tree_info.heading("No.", text="No.", anchor=tk.CENTER)
        tree_info.heading("台帳番號", text="台帳番號", anchor=tk.CENTER)
        tree_info.heading("內容", text="內容", anchor=tk.CENTER)
        tree_info.heading("業務狀況", text="業務狀況", anchor=tk.CENTER)
        tree_info.heading("備考", text="備考", anchor=tk.CENTER)
        tree_info.column("No.", anchor=tk.CENTER, width=15)
        tree_info.column("台帳番號", anchor=tk.CENTER)
        tree_info.column("內容", anchor=tk.CENTER)
        tree_info.column("業務狀況", anchor=tk.CENTER)
        tree_info.column("備考", anchor=tk.CENTER)
        tree_info.pack(fill='both', expand=True)

        # 標記樣式
        tree_info.tag_configure("highlight", background="lightblue")

        # 調整Treeview字體
        style = ttk.Style()
        style.configure("Treeview.Heading", font=font)
        style.configure("Treeview", font=font)

    root = tk.Toplevel()
    root.title("檢收日未登録")

    setup_ui(root)
    root.mainloop()

main_root = tk.Tk()
main_root.title("分室業務関連")

button_font = ("BIZ UDPゴシック", 15)

button1 = tk.Button(main_root, text="7日間通知書", command=run_program_1, font=button_font, width=15)
button1.pack(pady=10)

button2 = tk.Button(main_root, text="データ予定日早見表", command=program_1, font=button_font, width=15)
button2.pack(pady=10)

button3 = tk.Button(main_root, text="未発行図面", command=program_2, font=button_font, width=15)
button3.pack(pady=10)

# 新增的按鈕
button4 = tk.Button(main_root, text="台帳検収日未記入", command=program_3, font=button_font, width=15)
button4.pack(pady=10)

main_root.update_idletasks()
frame_width = main_root.winfo_reqwidth()
frame_height = main_root.winfo_reqheight()
main_root.geometry(f"{frame_width+20}x{frame_height+20}")

main_root.mainloop()
